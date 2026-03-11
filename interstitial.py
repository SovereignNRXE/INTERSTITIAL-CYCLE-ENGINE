"""
INTERSTITIAL CYCLE ENGINE
Autonomous Political Pressure Analysis // Self-Refining // Social Publishing
"""

import os
import json
import time
import datetime
import random
import re
import textwrap
import traceback
from pathlib import Path

import anthropic
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import numpy as np
from PIL import Image, ImageDraw, ImageFont
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from dotenv import load_dotenv

load_dotenv()

# ─────────────────────────────────────────────
# CONFIG
# ─────────────────────────────────────────────
OUTPUT_DIR   = Path(os.getenv("OUTPUT_DIR", "./output"))
DATA_FILE    = Path(os.getenv("DATA_FILE",  "./output/INTERSTITIAL_DATA.xlsx"))
LOGO_PATH    = Path(os.getenv("LOGO_PATH",  "./Interstitial_Logo.jpg"))
CYCLE_FILE   = Path(os.getenv("CYCLE_FILE", "./output/cycle_counter.txt"))
ENGINEERING_FILE = Path(os.getenv("ENGINEERING_FILE", "./output/ICE_ENGINEERING.xlsx"))
RESET_INTERVAL   = int(os.getenv("RESET_INTERVAL", "30"))  # cycles per subject before context reset
FONT_REG     = "/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf"
FONT_BOLD    = "/usr/share/fonts/truetype/liberation/LiberationSans-Bold.ttf"
FONT_FALLBACK = None  # will use matplotlib default if DejaVu missing

OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

# ─────────────────────────────────────────────
# CYCLE COUNTER — persistent, increments each run
# ─────────────────────────────────────────────

def get_cycle_number() -> int:
    if CYCLE_FILE.exists():
        try:
            return int(CYCLE_FILE.read_text().strip())
        except Exception:
            return 0
    return 0

def increment_cycle() -> int:
    n = get_cycle_number() + 1
    CYCLE_FILE.parent.mkdir(parents=True, exist_ok=True)
    CYCLE_FILE.write_text(str(n))
    return n

def format_cycle(n: int) -> str:
    return f"INTER//CYCLE {str(n).zfill(6)}"


# =============================================================================
# ALGORITHM 1 — REFINEMENT ENGINE
# Manages context window, 30-cycle reset, decay weighting
# =============================================================================

def subject_cycle_count(subject_name: str) -> int:
    """Count how many times this subject has appeared in RUNS sheet."""
    if not DATA_FILE.exists():
        return 0
    try:
        wb = openpyxl.load_workbook(DATA_FILE, data_only=True)
        ws = wb["RUNS"]
        return sum(1 for row in ws.iter_rows(min_row=2, values_only=True)
                   if row[2] == subject_name)
    except Exception:
        return 0

def needs_context_reset(subject_name: str) -> bool:
    """Return True if subject has hit the 30-cycle reset threshold."""
    count = subject_cycle_count(subject_name)
    return count > 0 and (count % RESET_INTERVAL == 0)

def load_refinement_context(subject_name: str = "") -> dict:
    """
    ALGORITHM 1 — REFINEMENT ENGINE
    Loads calibration context with decay weighting and reset logic.

    Returns dict with:
      - context_str: formatted string for prompt injection
      - reset_triggered: bool — True if 30-cycle reset fired
      - subject_appearances: int — how many times subject appeared in history
      - decay_weights: list of floats — weight assigned to each historical row
      - variance: float — score variance across subject history
    """
    result = {
        "context_str": "",
        "reset_triggered": False,
        "subject_appearances": 0,
        "decay_weights": [],
        "variance": 0.0,
        "anchor_strength": 0.0,
    }

    if not DATA_FILE.exists():
        return result

    try:
        wb    = openpyxl.load_workbook(DATA_FILE, data_only=True)
        ws    = wb["RUNS"]
        rows  = list(ws.iter_rows(min_row=2, values_only=True))
        if not rows:
            return result

        # Check reset threshold for this subject
        if subject_name:
            subject_rows = [r for r in rows if r[2] == subject_name]
            result["subject_appearances"] = len(subject_rows)

            if needs_context_reset(subject_name):
                result["reset_triggered"] = True
                log_engineering_event("CONTEXT_RESET", subject_name,
                                      result["subject_appearances"],
                                      {"reason": "30-cycle threshold reached",
                                       "action": "full context flush for subject"})
                # On reset: use only global recent context, exclude subject history
                recent = [r for r in rows[-20:] if r[2] != subject_name]
            else:
                recent = rows[-10:]
        else:
            recent = rows[-10:]

        if not recent:
            return result

        # Decay weighting: most recent rows get weight 1.0, oldest get 0.3
        n = len(recent)
        decay_weights = [0.3 + 0.7 * (i / max(n - 1, 1)) for i in range(n)]
        result["decay_weights"] = decay_weights

        # Compute anchor strength: how strongly prior scores dominate
        if subject_name:
            subj_scores = [r[24] for r in recent if r[2] == subject_name
                           and r[24] is not None]
            if len(subj_scores) >= 2:
                import statistics
                result["variance"]       = statistics.variance(subj_scores)
                result["anchor_strength"] = min(0.85, 0.3 + len(subj_scores) * 0.10)

        # Build prompt injection string with decay annotation
        lines = ["CALIBRATION CONTEXT (decay-weighted, newest = highest weight):"]
        for i, (row, w) in enumerate(zip(recent, decay_weights)):
            if row[2]:
                lines.append(
                    f"  [w={w:.2f}] {row[2]} | COMPOSITE: {row[24]}% "
                    f"| NODES: {row[26]} | {str(row[1])[:10]}"
                )

        if result["reset_triggered"]:
            lines.insert(1, f"  [RESET EVENT] Subject {subject_name} history flushed "
                            f"after {result['subject_appearances']} appearances. "
                            f"Recalibrate from first principles.")

        lines.append(
            "Drift instruction: Apply decay weights when anchoring to prior scores. "
            "Older entries (low w) should influence less. "
            "Avoid variance collapse -- scores must reflect genuine analytical position."
        )

        result["context_str"] = "\n".join(lines)
        return result

    except Exception as e:
        print(f"  [REFINEMENT] Context load error: {e}")
        return result


# =============================================================================
# ALGORITHM 2 — DIAGNOSTIC REPORTER
# Analyses performance metrics every N cycles, detects degradation
# =============================================================================

def run_diagnostic(cycle_n: int, current_data: dict) -> dict:
    """
    ALGORITHM 2 — DIAGNOSTIC REPORTER
    Fires every 30 cycles. Analyses:
      - Score variance across all recent runs (collapse detection)
      - Subject repeat frequency (pool saturation)
      - Composite mean drift (centrist regression)
      - Node score flatness per subject (anchoring failure)
      - Context reset events in window

    Returns dict with findings and HEALTH status.
    """
    diagnostic = {
        "cycle": cycle_n,
        "timestamp": datetime.datetime.utcnow().isoformat(),
        "health": "GREEN",
        "warnings": [],
        "metrics": {},
        "recommendation": "",
    }

    if not DATA_FILE.exists():
        diagnostic["health"] = "UNKNOWN"
        diagnostic["recommendation"] = "No data file yet. Run more cycles."
        return diagnostic

    try:
        wb   = openpyxl.load_workbook(DATA_FILE, data_only=True)
        ws   = wb["RUNS"]
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        if len(rows) < 5:
            diagnostic["health"] = "INSUFFICIENT_DATA"
            return diagnostic

        # Window: last 30 rows
        window = rows[-min(30, len(rows)):]
        composites = [r[24] for r in window if r[24] is not None]
        subjects   = [r[2]  for r in window if r[2]  is not None]

        if not composites:
            return diagnostic

        import statistics
        mean_comp = statistics.mean(composites)
        var_comp  = statistics.variance(composites) if len(composites) > 1 else 0
        stdev     = statistics.stdev(composites)    if len(composites) > 1 else 0

        diagnostic["metrics"] = {
            "window_size":     len(window),
            "composite_mean":  round(mean_comp, 1),
            "composite_stdev": round(stdev, 1),
            "composite_variance": round(var_comp, 1),
            "subject_unique":  len(set(subjects)),
            "subject_total":   len(subjects),
        }

        # ── VARIANCE COLLAPSE check ──────────────────────────────────────────
        if var_comp < 25:
            diagnostic["warnings"].append(
                f"VARIANCE COLLAPSE: composite variance={var_comp:.1f} (threshold <25). "
                f"Scores converging on mean {mean_comp:.1f}%. Analytical discrimination failing."
            )
            diagnostic["health"] = "AMBER"

        if var_comp < 10:
            diagnostic["warnings"].append(
                "CRITICAL VARIANCE COLLAPSE: engine in hall-of-mirrors state. "
                "Context reset recommended immediately."
            )
            diagnostic["health"] = "RED"

        # ── CENTRIST REGRESSION check ────────────────────────────────────────
        if 48 <= mean_comp <= 57 and stdev < 8:
            diagnostic["warnings"].append(
                f"CENTRIST REGRESSION: mean {mean_comp:.1f}%, stdev {stdev:.1f}. "
                f"Model defaulting to safe mid-range scores. Prompt differentiation failing."
            )
            if diagnostic["health"] == "GREEN":
                diagnostic["health"] = "AMBER"

        # ── POOL SATURATION check ────────────────────────────────────────────
        saturation = len(set(subjects)) / max(len(subjects), 1)
        diagnostic["metrics"]["pool_saturation"] = round(saturation, 2)
        if saturation < 0.5:
            diagnostic["warnings"].append(
                f"POOL SATURATION: only {len(set(subjects))} unique subjects "
                f"in last {len(subjects)} runs (ratio {saturation:.2f}). "
                f"Subject rotation insufficient."
            )

        # ── NODE FLATNESS per subject ────────────────────────────────────────
        # Check if any subject shows <3pt stdev across all 6 node scores
        subject_node_scores = {}
        for r in window:
            subj = r[2]
            if not subj:
                continue
            node_cols = [r[7], r[10], r[13], r[16], r[19], r[22]]  # N01-N06 scores
            scores = [s for s in node_cols if s is not None]
            if subj not in subject_node_scores:
                subject_node_scores[subj] = []
            subject_node_scores[subj].extend(scores)

        flat_subjects = []
        for subj, scores in subject_node_scores.items():
            if len(scores) >= 6:
                sv = statistics.stdev(scores) if len(scores) > 1 else 0
                if sv < 4:
                    flat_subjects.append(f"{subj}(σ={sv:.1f})")

        if flat_subjects:
            diagnostic["warnings"].append(
                f"NODE FLATNESS detected: {', '.join(flat_subjects)}. "
                f"Node scores too uniform — subject-specific differentiation lost."
            )
            if diagnostic["health"] == "GREEN":
                diagnostic["health"] = "AMBER"

        # ── RECOMMENDATION ───────────────────────────────────────────────────
        if diagnostic["health"] == "GREEN":
            diagnostic["recommendation"] = (
                f"ENGINE NOMINAL. Variance {var_comp:.1f}, mean {mean_comp:.1f}%, "
                f"pool utilisation {saturation:.0%}. Continue."
            )
        elif diagnostic["health"] == "AMBER":
            diagnostic["recommendation"] = (
                "DEGRADATION DETECTED. Monitor next 10 cycles. "
                "If variance continues falling, flush context for affected subjects."
            )
        else:
            diagnostic["recommendation"] = (
                "INTERVENTION REQUIRED. Force context reset for all subjects "
                "appearing 3+ times in last 30 cycles. Consider prompt revision."
            )

        # Log to engineering store
        log_engineering_event("DIAGNOSTIC", "SYSTEM", cycle_n, diagnostic)
        return diagnostic

    except Exception as e:
        diagnostic["health"] = "ERROR"
        diagnostic["warnings"].append(str(e))
        return diagnostic


# =============================================================================
# ALGORITHM 3 — ENGINEERING DATA STORE
# Writes all system events to ICE_ENGINEERING.xlsx (separate from content data)
# =============================================================================

def init_engineering_store():
    """Initialise ICE_ENGINEERING.xlsx if it does not exist."""
    if ENGINEERING_FILE.exists():
        return
    wb = openpyxl.Workbook()

    # Sheet 1: EVENTS — one row per engineering event
    ws = wb.active
    ws.title = "EVENTS"
    headers = [
        "EVENT_ID", "EVENT_TYPE", "TIMESTAMP", "CYCLE_NUMBER",
        "SUBJECT", "HEALTH", "DETAIL_JSON"
    ]
    hf = Font(bold=True, color="FFFFFF")
    hfill = PatternFill("solid", start_color="0A0A1A")
    ws.append(headers)
    for cell in ws[1]:
        cell.font  = hf
        cell.fill  = hfill
        cell.alignment = Alignment(horizontal='center')
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 25

    # Sheet 2: DIAGNOSTICS — structured diagnostic snapshots
    ws2 = wb.create_sheet("DIAGNOSTICS")
    diag_headers = [
        "DIAG_ID", "TIMESTAMP", "CYCLE", "HEALTH",
        "COMPOSITE_MEAN", "COMPOSITE_STDEV", "COMPOSITE_VARIANCE",
        "SUBJECT_UNIQUE", "POOL_SATURATION",
        "WARNINGS_COUNT", "WARNINGS_TEXT", "RECOMMENDATION"
    ]
    ws2.append(diag_headers)
    for cell in ws2[1]:
        cell.font  = hf
        cell.fill  = hfill
        cell.alignment = Alignment(horizontal='center')
    for col in ws2.columns:
        ws2.column_dimensions[col[0].column_letter].width = 22

    # Sheet 3: RESETS — context reset events
    ws3 = wb.create_sheet("RESETS")
    reset_headers = [
        "RESET_ID", "TIMESTAMP", "CYCLE", "SUBJECT",
        "APPEARANCES_AT_RESET", "REASON"
    ]
    ws3.append(reset_headers)
    for cell in ws3[1]:
        cell.font  = hf
        cell.fill  = hfill
        cell.alignment = Alignment(horizontal='center')
    for col in ws3.columns:
        ws3.column_dimensions[col[0].column_letter].width = 25

    # Sheet 4: SUMMARY — live formula metrics
    ws4 = wb.create_sheet("SUMMARY")
    ws4.append(["METRIC", "VALUE", "SOURCE"])
    ws4.append(["Total events",      "=COUNTA(EVENTS!A:A)-1",          "EVENTS sheet"])
    ws4.append(["Total diagnostics", "=COUNTA(DIAGNOSTICS!A:A)-1",     "DIAGNOSTICS sheet"])
    ws4.append(["Total resets",      "=COUNTA(RESETS!A:A)-1",          "RESETS sheet"])
    ws4.append(["RED health count",  '=COUNTIF(DIAGNOSTICS!D:D,"RED")', "DIAGNOSTICS sheet"])
    ws4.append(["AMBER health count",'=COUNTIF(DIAGNOSTICS!D:D,"AMBER")', "DIAGNOSTICS sheet"])
    ws4.append(["Avg composite mean","=AVERAGE(DIAGNOSTICS!E:E)",       "DIAGNOSTICS sheet"])
    ws4.append(["Avg variance",      "=AVERAGE(DIAGNOSTICS!G:G)",       "DIAGNOSTICS sheet"])
    ws4.append(["Min variance seen", "=MIN(DIAGNOSTICS!G:G)",           "Collapse indicator"])
    for cell in ws4[1]:
        cell.font = Font(bold=True)
    for col in ws4.columns:
        ws4.column_dimensions[col[0].column_letter].width = 28

    wb.save(str(ENGINEERING_FILE))
    print(f"  [ENG] Initialised {ENGINEERING_FILE}")


def log_engineering_event(event_type: str, subject: str,
                           cycle_or_appearances, detail: dict):
    """Write one row to EVENTS sheet and structured row to relevant sheet."""
    init_engineering_store()
    try:
        wb = openpyxl.load_workbook(str(ENGINEERING_FILE))
        now = datetime.datetime.utcnow().isoformat()
        event_id = f"EVT-{now[:19].replace(':','').replace('-','').replace('T','-')}"
        health   = detail.get("health", "")

        # EVENTS sheet — always
        ws = wb["EVENTS"]
        ws.append([
            event_id, event_type, now, cycle_or_appearances,
            subject, health, json.dumps(detail)[:2000]
        ])
        # Colour health column
        last = ws.max_row
        hcol = ws.cell(row=last, column=6)
        if health == "RED":
            hcol.fill = PatternFill("solid", start_color="4A0A0A")
        elif health == "AMBER":
            hcol.fill = PatternFill("solid", start_color="3A2A00")
        elif health == "GREEN":
            hcol.fill = PatternFill("solid", start_color="0A2A0A")

        # DIAGNOSTICS sheet
        if event_type == "DIAGNOSTIC":
            ws2 = wb["DIAGNOSTICS"]
            m   = detail.get("metrics", {})
            ws2.append([
                event_id, now,
                detail.get("cycle", ""),
                health,
                m.get("composite_mean", ""),
                m.get("composite_stdev", ""),
                m.get("composite_variance", ""),
                m.get("subject_unique", ""),
                m.get("pool_saturation", ""),
                len(detail.get("warnings", [])),
                " | ".join(detail.get("warnings", [])),
                detail.get("recommendation", ""),
            ])

        # RESETS sheet
        if event_type == "CONTEXT_RESET":
            ws3 = wb["RESETS"]
            ws3.append([
                event_id, now, cycle_or_appearances, subject,
                detail.get("subject_appearances",
                           cycle_or_appearances),
                detail.get("reason", "30-cycle threshold")
            ])

        wb.save(str(ENGINEERING_FILE))
    except Exception as e:
        print(f"  [ENG] Log error: {e}")


# Social media credentials (set in .env)
# Mastodon
MASTODON_ACCESS_TOKEN = os.getenv("MASTODON_ACCESS_TOKEN", "")
MASTODON_INSTANCE     = os.getenv("MASTODON_INSTANCE", "")   # e.g. mastodon.social

# Bluesky
BSKY_HANDLE           = os.getenv("BSKY_HANDLE", "")         # e.g. yourhandle.bsky.social
BSKY_APP_PASSWORD     = os.getenv("BSKY_APP_PASSWORD", "")   # App password from bsky settings

# Telegram
TELEGRAM_BOT_TOKEN    = os.getenv("TELEGRAM_BOT_TOKEN", "")  # From @BotFather
TELEGRAM_CHANNEL_ID   = os.getenv("TELEGRAM_CHANNEL_ID", "") # e.g. @yourchannel or -1001234567890

# LinkedIn
LINKEDIN_ACCESS_TOKEN = os.getenv("LINKEDIN_ACCESS_TOKEN", "")
LINKEDIN_PERSON_URN   = os.getenv("LINKEDIN_PERSON_URN", "") # e.g. urn:li:person:XXXXXXXX

# ─────────────────────────────────────────────
# SUBJECT POOL
# ─────────────────────────────────────────────
SUBJECTS = [
    {"name": "UNITED STATES", "type": "COUNTRY", "scope": "DOMESTIC COHESION AND GLOBAL HEGEMONY RETENTION THROUGH 2028"},
    {"name": "CHINA", "type": "COUNTRY", "scope": "CCP LEGITIMACY, TAIWAN POLICY TRAJECTORY, AND ECONOMIC STABILISATION"},
    {"name": "RUSSIA", "type": "COUNTRY", "scope": "KREMLIN POWER RETENTION, WAR ECONOMY SUSTAINABILITY, AND WESTERN PRESSURE TOLERANCE"},
    {"name": "GERMANY", "type": "COUNTRY", "scope": "COALITION STABILITY, INDUSTRIAL TRANSITION, AND EU LEADERSHIP CAPACITY"},
    {"name": "FRANCE", "type": "COUNTRY", "scope": "FIFTH REPUBLIC AUTHORITY, SAHEL WITHDRAWAL FALLOUT, AND DOMESTIC POLARISATION"},
    {"name": "UNITED KINGDOM", "type": "COUNTRY", "scope": "POST-BREXIT ECONOMIC RECOVERY, DEVOLUTION PRESSURE, AND DEFENCE POSTURE"},
    {"name": "JAPAN", "type": "COUNTRY", "scope": "CONSTITUTIONAL REVISION TRAJECTORY, CHINA DETERRENCE, AND ECONOMIC STAGNATION"},
    {"name": "INDIA", "type": "COUNTRY", "scope": "BJP COALITION DURABILITY, BORDER TENSION MANAGEMENT, AND GREAT POWER POSITIONING"},
    {"name": "BRAZIL", "type": "COUNTRY", "scope": "LULA COALITION COHESION, AMAZON POLICY, AND REGIONAL LEADERSHIP CREDIBILITY"},
    {"name": "CANADA", "type": "COUNTRY", "scope": "FEDERAL COHESION, US TARIFF EXPOSURE, AND IMMIGRATION STRAIN"},
    {"name": "ITALY", "type": "COUNTRY", "scope": "MELONI COALITION DURABILITY, EU FISCAL COMPLIANCE, AND MIGRATION PRESSURE"},
    {"name": "SOUTH KOREA", "type": "COUNTRY", "scope": "PRESIDENTIAL AUTHORITY POST-MARTIAL LAW, NORTH KOREA DETERRENCE, US ALLIANCE"},
    {"name": "AUSTRALIA", "type": "COUNTRY", "scope": "AUKUS DELIVERY CAPACITY, CHINA TRADE DEPENDENCY, AND DOMESTIC ENERGY TRANSITION"},
    {"name": "SPAIN", "type": "COUNTRY", "scope": "SANCHEZ MINORITY GOVERNMENT SURVIVAL, CATALAN SETTLEMENT, AND ECONOMIC TRAJECTORY"},
    {"name": "NETHERLANDS", "type": "COUNTRY", "scope": "WILDERS COALITION STABILITY, EU MIGRATION POLICY INFLUENCE, AND DEFENCE SPENDING"},
    {"name": "SAUDI ARABIA", "type": "COUNTRY", "scope": "MBS VISION 2030 DELIVERY, OIL PRICE DEPENDENCY, AND REGIONAL HEGEMON COMPETITION"},
    {"name": "IRAN", "type": "COUNTRY", "scope": "REGIME LEGITIMACY POST-MAHSA, NUCLEAR PROGRAMME TRAJECTORY, AND SANCTIONS PRESSURE"},
    {"name": "ISRAEL", "type": "COUNTRY", "scope": "WAR CABINET COHESION, GAZA OPERATION SUSTAINABILITY, AND REGIONAL NORMALISATION"},
    {"name": "TURKEY", "type": "COUNTRY", "scope": "ERDOGAN CONSOLIDATION, INFLATION CRISIS, AND NATO-RUSSIA BALANCING"},
    {"name": "INDONESIA", "type": "COUNTRY", "scope": "PRABOWO TRANSITION, SOUTH CHINA SEA EXPOSURE, AND COMMODITY DEPENDENCY"},
    {"name": "MEXICO", "type": "COUNTRY", "scope": "CLAUDIA SHEINBAUM AUTHORITY, CARTEL TERRITORIAL CONTROL, AND US RELATIONS"},
    {"name": "ARGENTINA", "type": "COUNTRY", "scope": "MILEI SHOCK THERAPY SURVIVAL, IMF COMPLIANCE, AND SOCIAL COHESION"},
    {"name": "EGYPT", "type": "COUNTRY", "scope": "SISI REGIME DURABILITY, IMF DEPENDENCY, AND REGIONAL INSTABILITY EXPOSURE"},
    {"name": "PAKISTAN", "type": "COUNTRY", "scope": "CIVIL-MILITARY BALANCE, ECONOMIC COLLAPSE RISK, AND AFGHANISTAN BORDER PRESSURE"},
    {"name": "NIGERIA", "type": "COUNTRY", "scope": "TINUBU AUTHORITY, SUBSIDY REMOVAL FALLOUT, AND BOKO HARAM TERRITORIAL CONTROL"},
    {"name": "UKRAINE", "type": "COUNTRY", "scope": "WAR SUSTAINABILITY, TERRITORIAL CONTROL TRAJECTORY, AND WESTERN SUPPORT DEPENDENCE"},
    {"name": "POLAND", "type": "COUNTRY", "scope": "TUSK COALITION RESILIENCE, EASTERN FLANK SECURITY ROLE, AND EU LAW REHABILITATION"},
    {"name": "HUNGARY", "type": "COUNTRY", "scope": "ORBAN EU LEVERAGE, ECONOMIC EXPOSURE, AND DEMOCRATIC BACKSLIDING TRAJECTORY"},
    {"name": "SERBIA", "type": "COUNTRY", "scope": "VUCIC BALANCING BETWEEN EU AND RUSSIA, KOSOVO STATUS, AND OPPOSITION PRESSURE"},
    {"name": "ETHIOPIA", "type": "COUNTRY", "scope": "ABIY AUTHORITY POST-TIGRAY, AMHARA INSURGENCY, AND NILE DISPUTE WITH EGYPT"},
    {"name": "MYANMAR", "type": "COUNTRY", "scope": "JUNTA TERRITORIAL CONTROL, RESISTANCE COALITION STRENGTH, AND HUMANITARIAN COLLAPSE"},
    {"name": "VENEZUELA", "type": "COUNTRY", "scope": "MADURO LEGITIMACY POST-ELECTION FRAUD, OPPOSITION RESILIENCE, AND OIL REVENUE"},
    {"name": "COLOMBIA", "type": "COUNTRY", "scope": "PETRO REFORM AGENDA VIABILITY, FARC DISSIDENT RESURGENCE, AND US RELATIONS"},
    {"name": "SOUTH AFRICA", "type": "COUNTRY", "scope": "ANC-LED GNU COHESION, LOAD SHEDDING CRISIS, AND BRICS POSITIONING"},
    {"name": "KENYA", "type": "COUNTRY", "scope": "RUTO AUTHORITY POST-PROTESTS, IMF COMPLIANCE, AND EAST AFRICA STABILITY ROLE"},
    {"name": "SUDAN", "type": "COUNTRY", "scope": "RSF VS SAF WAR TRAJECTORY, HUMANITARIAN COLLAPSE, AND INTERNATIONAL INTERVENTION"},
    {"name": "SOMALIA", "type": "COUNTRY", "scope": "FEDERAL GOVERNMENT REACH, AL-SHABAAB TERRITORIAL CONTROL, AND DROUGHT PRESSURE"},
    {"name": "MALI", "type": "COUNTRY", "scope": "JUNTA-WAGNER RELATIONSHIP POST-PRIGOZHIN, FRENCH WITHDRAWAL FALLOUT, AND JIHADIST EXPANSION"},
    {"name": "NIGER", "type": "COUNTRY", "scope": "JUNTA CONSOLIDATION, ECOWAS PRESSURE TOLERANCE, AND URANIUM LEVERAGE"},
    {"name": "HAITI", "type": "COUNTRY", "scope": "GANG TERRITORIAL CONTROL, TRANSITIONAL COUNCIL VIABILITY, AND KENYAN MISSION CAPACITY"},
    {"name": "GEORGIA", "type": "COUNTRY", "scope": "GEORGIAN DREAM EU PIVOT REVERSAL, PROTEST MOVEMENT DURABILITY, AND RUSSIAN PRESSURE"},
    {"name": "MOLDOVA", "type": "COUNTRY", "scope": "EU ACCESSION TRAJECTORY, TRANSNISTRIA DEPENDENCY LEVERAGE, AND RUSSIAN HYBRID PRESSURE"},
    {"name": "ARMENIA", "type": "COUNTRY", "scope": "POST-KARABAKH SOVEREIGNTY REORIENTATION, CSTO WITHDRAWAL, AND EU ALIGNMENT"},
    {"name": "AZERBAIJAN", "type": "COUNTRY", "scope": "ALIYEV CONSOLIDATION POST-KARABAKH, ENERGY LEVERAGE, AND REGIONAL HEGEMON POSTURE"},
    {"name": "KAZAKHSTAN", "type": "COUNTRY", "scope": "TOKAYEV AUTHORITY POST-2022 PROTESTS, RUSSIA DEPENDENCY, AND WESTERN PIVOT LIMITS"},
    {"name": "AFGHANISTAN", "type": "COUNTRY", "scope": "TALIBAN GOVERNANCE CAPACITY, INTERNATIONAL ISOLATION, AND HUMANITARIAN COLLAPSE"},
    {"name": "IRAQ", "type": "COUNTRY", "scope": "SHIA BLOC COHESION, US TROOP PRESSURE, AND IRANIAN PROXY NETWORK INFLUENCE"},
    {"name": "SYRIA", "type": "COUNTRY", "scope": "POST-ASSAD TRANSITION FRAGILITY, TERRITORIAL CONTROL PATCHWORK, AND RECONSTRUCTION"},
    {"name": "LIBYA", "type": "COUNTRY", "scope": "GNU VS GNA STALEMATE, FOREIGN PROXY WAR DYNAMICS, AND OIL REVENUE CAPTURE"},
    {"name": "LEBANON", "type": "COUNTRY", "scope": "POST-WAR RECONSTRUCTION CAPACITY, HEZBOLLAH DISARMAMENT, AND BANKING COLLAPSE"},
    {"name": "YEMEN", "type": "COUNTRY", "scope": "HOUTHI RED SEA STRATEGY, UN PEACE PROCESS, AND HUMANITARIAN CATASTROPHE"},
    {"name": "BANGLADESH", "type": "COUNTRY", "scope": "POST-HASINA TRANSITION, STUDENT MOVEMENT INSTITUTIONALISATION, AND IMF DEPENDENCY"},
    {"name": "SRI LANKA", "type": "COUNTRY", "scope": "DISSANAYAKE REFORM AGENDA, IMF PROGRAMME COMPLIANCE, AND DEBT RESTRUCTURING"},
    {"name": "ZIMBABWE", "type": "COUNTRY", "scope": "MNANGAGWA AUTHORITY, CURRENCY COLLAPSE CYCLE, AND OPPOSITION SUPPRESSION"},
    {"name": "CUBA", "type": "COUNTRY", "scope": "DIAZ-CANEL LEGITIMACY, POWER GRID COLLAPSE, AND US SANCTIONS PRESSURE"},
    {"name": "NICARAGUA", "type": "COUNTRY", "scope": "ORTEGA-MURILLO DYNASTIC CONSOLIDATION, CHURCH SUPPRESSION, AND ISOLATION"},
    {"name": "THAILAND", "type": "COUNTRY", "scope": "PAETONGTARN GOVERNMENT STABILITY, MILITARY OVERSIGHT, AND CONSTITUTIONAL REFORM"},
    {"name": "MALAYSIA", "type": "COUNTRY", "scope": "ANWAR COALITION DURABILITY, CHINA TRADE DEPENDENCY, AND ISLAMIC CONSERVATISM PRESSURE"},
    {"name": "PHILIPPINES", "type": "COUNTRY", "scope": "MARCOS-DUTERTE COALITION FRACTURE, SOUTH CHINA SEA CONFRONTATIONS, AND US ALLIANCE"},
    {"name": "VIETNAM", "type": "COUNTRY", "scope": "CPV LEADERSHIP POST-TO LAM TRANSITION, US-CHINA BALANCING, AND FDI DEPENDENCY"},
    {"name": "CAMBODIA", "type": "COUNTRY", "scope": "HUN MANET DYNASTIC SUCCESSION, CHINESE DEBT DEPENDENCY, AND OPPOSITION SUPPRESSION"},
    {"name": "NORTH KOREA", "type": "COUNTRY", "scope": "KIM JONG-UN CONSOLIDATION, NUCLEAR PROGRAMME TRAJECTORY, AND RUSSIA ALIGNMENT"},
    {"name": "NEPAL", "type": "COUNTRY", "scope": "COALITION GOVERNMENT FRAGILITY, CHINA-INDIA BALANCING, AND REMITTANCE DEPENDENCY"},
    {"name": "BOLIVIA", "type": "COUNTRY", "scope": "ARC SPLIT, GAS REVENUE DECLINE, AND COUP ATTEMPT AFTERMATH"},
    {"name": "ECUADOR", "type": "COUNTRY", "scope": "NOBOA SECURITY OFFENSIVE SUSTAINABILITY, GANG TERRITORIAL CONTROL, AND OIL DECLINE"},
    {"name": "PERU", "type": "COUNTRY", "scope": "BOLUARTE AUTHORITY, CONGRESSIONAL DYSFUNCTION, AND MINING SECTOR CONFLICT"},
    {"name": "CHILE", "type": "COUNTRY", "scope": "BORIC REFORM RETREAT, CONSTITUTIONAL PROCESS FAILURE AFTERMATH, AND CRIME SURGE"},
    {"name": "SENEGAL", "type": "COUNTRY", "scope": "FAYE NEW GOVERNMENT CONSOLIDATION, IMF RENEGOTIATION, AND WEST AFRICA STABILITY ROLE"},
    {"name": "GHANA", "type": "COUNTRY", "scope": "MAHAMA RETURN, DEBT RESTRUCTURING COMPLETION, AND COCOA SECTOR RECOVERY"},
    {"name": "TANZANIA", "type": "COUNTRY", "scope": "HASSAN AUTHORITY, ZANZIBAR AUTONOMY PRESSURE, AND EAST AFRICA PIVOT"},
    {"name": "MOZAMBIQUE", "type": "COUNTRY", "scope": "POST-ELECTION VIOLENCE TRAJECTORY, FRELIMO LEGITIMACY, AND CABO DELGADO INSURGENCY"},
    {"name": "DRC", "type": "COUNTRY", "scope": "TSHISEKEDI TERRITORIAL CONTROL, M23 ADVANCE, AND MINERAL RESOURCE CONFLICT"},
    {"name": "CAMEROON", "type": "COUNTRY", "scope": "BIYA SUCCESSION UNCERTAINTY, ANGLOPHONE SEPARATIST CONFLICT, AND SAHEL EXPOSURE"},
    {"name": "ANGOLA", "type": "COUNTRY", "scope": "LOURENCO REFORM TRAJECTORY, OIL DEPENDENCY, AND CHINA DEBT LEVERAGE"},
    {"name": "ZAMBIA", "type": "COUNTRY", "scope": "HICHILEMA DEBT RESTRUCTURING, COPPER SECTOR DEPENDENCY, AND DEMOCRATIC CONSOLIDATION"},
    {"name": "TUNISIA", "type": "COUNTRY", "scope": "SAIED PRESIDENTIAL CONSOLIDATION, IMF PROGRAMME COLLAPSE RISK, AND MIGRATION PRESSURE"},
    {"name": "MOROCCO", "type": "COUNTRY", "scope": "MOHAMMED VI AUTHORITY, SAHARA RECOGNITION MOMENTUM, AND MIGRATION LEVERAGE"},
    {"name": "ALGERIA", "type": "COUNTRY", "scope": "TEBBOUNE CONSOLIDATION, GAS EXPORT LEVERAGE, AND DOMESTIC REFORM LIMITS"},
    {"name": "JORDAN", "type": "COUNTRY", "scope": "ABDULLAH STABILITY, GAZA WAR REFUGEE PRESSURE, AND US AID DEPENDENCY"},
    {"name": "OMAN", "type": "COUNTRY", "scope": "HAITHAM CONSOLIDATION, IRAN MEDIATION ROLE, AND OIL REVENUE DIVERSIFICATION"},
    {"name": "QATAR", "type": "COUNTRY", "scope": "AL-THANI REGIONAL MEDIATION CAPACITY, LNG LEVERAGE, AND WORLD CUP LEGACY"},
    {"name": "UAE", "type": "COUNTRY", "scope": "AL-NAHYAN DIVERSIFICATION TRAJECTORY, ABRAHAM ACCORDS EXPANSION, AND CHINA BALANCE"},
    {"name": "UZBEKISTAN", "type": "COUNTRY", "scope": "MIRZIYOYEV REFORM PACE, RUSSIA DEPENDENCY REDUCTION, AND REGIONAL CONNECTIVITY"},
    {"name": "TAJIKISTAN", "type": "COUNTRY", "scope": "RAHMON DYNASTIC SUCCESSION, AFGHANISTAN BORDER PRESSURE, AND RUSSIA DEPENDENCY"},
    {"name": "KYRGYZSTAN", "type": "COUNTRY", "scope": "JAPAROV AUTHORITY, RUSSIA MIGRANT REMITTANCE DEPENDENCY, AND ETHNIC TENSION"},
    {"name": "TURKMENISTAN", "type": "COUNTRY", "scope": "BERDIMUHAMEDOW DYNASTIC CONSOLIDATION, GAS EXPORT ISOLATION, AND HUMAN RIGHTS"},
    {"name": "MONGOLIA", "type": "COUNTRY", "scope": "DEMOCRATIC CONSOLIDATION, CHINA-RUSSIA SANDWICH LEVERAGE, AND MINING DEPENDENCY"},
    {"name": "PAPUA NEW GUINEA", "type": "COUNTRY", "scope": "MARAPE AUTHORITY, RESOURCE NATIONALISM, AND AUSTRALIA-CHINA COMPETITION"},
    {"name": "FIJI", "type": "COUNTRY", "scope": "RABUKA COALITION STABILITY, CLIMATE VULNERABILITY, AND CHINA INFLUENCE RESISTANCE"},
    {"name": "TIMOR-LESTE", "type": "COUNTRY", "scope": "RAMOS-HORTA AUTHORITY, OIL REVENUE DEPLETION, AND ASEAN ACCESSION TRAJECTORY"},
    {"name": "LAOS", "type": "COUNTRY", "scope": "LPRP AUTHORITY, CHINESE DEBT TRAP DEPTH, AND MEKONG HYDROPOWER STRATEGY"},
    {"name": "BRUNEI", "type": "COUNTRY", "scope": "HASSANAL BOLKIAH ABSOLUTE RULE, OIL DEPLETION TIMELINE, AND SHARIA LAW SCOPE"},
    {"name": "BHUTAN", "type": "COUNTRY", "scope": "WANGCHUCK MONARCHY STABILITY, INDIA DEPENDENCY, AND CHINA BORDER NEGOTIATION"},
    {"name": "MALDIVES", "type": "COUNTRY", "scope": "MUIZZU INDIA-OUT POLICY SUSTAINABILITY, CLIMATE EXISTENTIAL RISK, AND CHINA DEBT"},
    {"name": "ICELAND", "type": "COUNTRY", "scope": "ARCTIC SOVEREIGNTY, NATO ROLE, AND FISHERIES DISPUTE MANAGEMENT"},
    {"name": "IRELAND", "type": "COUNTRY", "scope": "EU FISCAL SOVEREIGNTY, US TECH DEPENDENCY, AND NORTHERN IRELAND PROTOCOL"},
    {"name": "PORTUGAL", "type": "COUNTRY", "scope": "DEMOCRATIC STABILITY, GOLDEN VISA REFORM, AND ATLANTIC GATEWAY ROLE"},
    {"name": "GREECE", "type": "COUNTRY", "scope": "DEBT RECOVERY TRAJECTORY, MIGRATION PRESSURE, AND AEGEAN SOVEREIGNTY DISPUTES"},
    {"name": "CZECHIA", "type": "COUNTRY", "scope": "FIALA COALITION DURABILITY, ENERGY TRANSITION, AND V4 POSITIONING"},
    {"name": "SLOVAKIA", "type": "COUNTRY", "scope": "FICO RUSSIA ALIGNMENT, EU COHESION FUND LEVERAGE, AND NATO CREDIBILITY"},
    {"name": "ROMANIA", "type": "COUNTRY", "scope": "POST-ELECTION INSTABILITY, NATO EASTERN FLANK ROLE, AND SCHENGEN INTEGRATION"},
    {"name": "BULGARIA", "type": "COUNTRY", "scope": "POLITICAL PARALYSIS CYCLE, CORRUPTION REFORM, AND SCHENGEN ENTRY"},
    {"name": "CROATIA", "type": "COUNTRY", "scope": "EURO ZONE INTEGRATION CONSOLIDATION, BALKANS STABILITY ROLE, AND MIGRATION ROUTE"},
    {"name": "SLOVENIA", "type": "COUNTRY", "scope": "GOLOB AUTHORITY, EU PRESIDENCY LEGACY, AND BALKANS BRIDGE ROLE"},
    {"name": "AUSTRIA", "type": "COUNTRY", "scope": "FPO COALITION FORMATION, NEUTRALITY POSTURE, AND MIGRATION POLICY"},
    {"name": "SWITZERLAND", "type": "COUNTRY", "scope": "BILATERAL AGREEMENT WITH EU, BANKING SECTOR STABILITY, AND NEUTRALITY EROSION"},
    {"name": "BELGIUM", "type": "COUNTRY", "scope": "COALITION FORMATION RECORD, EU CAPITAL STATUS, AND COMMUNITY TENSION"},
    {"name": "LUXEMBOURG", "type": "COUNTRY", "scope": "FINANCIAL CENTRE REGULATION PRESSURE, EU FISCAL POLICY INFLUENCE, AND NATO CONTRIBUTION"},
    {"name": "DENMARK", "type": "COUNTRY", "scope": "FREDERIKSEN AUTHORITY, GREENLAND SOVEREIGNTY PRESSURE, AND DEFENCE SPENDING"},
    {"name": "SWEDEN", "type": "COUNTRY", "scope": "NATO INTEGRATION, GANG VIOLENCE CRISIS, AND KRISTERSSON COALITION"},
    {"name": "NORWAY", "type": "COUNTRY", "scope": "OIL FUND SOVEREIGN WEALTH STRATEGY, NATO ARCTIC ROLE, AND STORTING STABILITY"},
    {"name": "FINLAND", "type": "COUNTRY", "scope": "NATO EASTERN FLANK ROLE, RUSSIA BORDER SECURITY, AND ORPO COALITION"},
    {"name": "ESTONIA", "type": "COUNTRY", "scope": "NATO FRONTLINE EXPOSURE, DIGITAL GOVERNANCE LEADERSHIP, AND RUSSIA HYBRID THREAT"},
    {"name": "LATVIA", "type": "COUNTRY", "scope": "NATO FRONTLINE EXPOSURE, RUSSIAN MINORITY INTEGRATION, AND ENERGY INDEPENDENCE"},
    {"name": "LITHUANIA", "type": "COUNTRY", "scope": "SUWALKI CORRIDOR VULNERABILITY, KALININGRAD PRESSURE, AND NATO COMMITMENT"},
    {"name": "ALBANIA", "type": "COUNTRY", "scope": "EU ACCESSION CHAPTER PROGRESS, RAMA AUTHORITY, AND ORGANISED CRIME REFORM"},
    {"name": "NORTH MACEDONIA", "type": "COUNTRY", "scope": "EU ACCESSION STALL, IDENTITY DISPUTE LEGACY, AND BULGARIAN VETO"},
    {"name": "BOSNIA", "type": "COUNTRY", "scope": "REPUBLIKA SRPSKA SECESSION THREAT, EU ACCESSION, AND DAYTON FRAMEWORK EROSION"},
    {"name": "KOSOVO", "type": "COUNTRY", "scope": "KURTI AUTHORITY, RECOGNITION EXPANSION, AND SERBIA NORMALISATION"},
    {"name": "MONTENEGRO", "type": "COUNTRY", "scope": "NATO MEMBER CONSOLIDATION, SERBIAN CHURCH DISPUTE, AND EU ACCESSION"},
    {"name": "CYPRUS", "type": "COUNTRY", "scope": "REUNIFICATION PROCESS, ENERGY FIELD DISPUTE, AND EU SOVEREIGNTY PROTECTION"},
    {"name": "MALTA", "type": "COUNTRY", "scope": "ABELA AUTHORITY, MIGRATION BURDEN, AND FINANCIAL PASSPORT SCHEME LEGACY"},
    {"name": "LIECHTENSTEIN", "type": "COUNTRY", "scope": "FINANCIAL SOVEREIGNTY, EEA DEPENDENCY, AND BANKING INTEGRITY REGIME"},
    {"name": "ANDORRA", "type": "COUNTRY", "scope": "EU ASSOCIATION AGREEMENT NEGOTIATION, TOURISM DEPENDENCY, AND TAX HAVEN STATUS"},
    {"name": "MONACO", "type": "COUNTRY", "scope": "GRIMALDI DYNASTY STABILITY, FISCAL TREATY WITH FRANCE, AND SUPER-WEALTH CONCENTRATION"},
    {"name": "SAN MARINO", "type": "COUNTRY", "scope": "EU ASSOCIATION TALKS, INSTITUTIONAL REFORM, AND BANKING TRANSPARENCY PRESSURE"},
    {"name": "VATICAN", "type": "COUNTRY", "scope": "PAPAL AUTHORITY POST-FRANCIS, DIPLOMATIC REACH, AND DOCTRINAL REFORM RESISTANCE"},
    {"name": "NEW ZEALAND", "type": "COUNTRY", "scope": "LUXON GOVERNMENT REFORM AGENDA, MAORI TREATY CONFLICT, AND PACIFIC SECURITY ROLE"},
    {"name": "SINGAPORE", "type": "COUNTRY", "scope": "PAP DOMINANCE POST-SUCCESSION, CHINA-US BALANCING, AND FINANCIAL CENTRE RESILIENCE"},
    {"name": "TAIWAN", "type": "COUNTRY", "scope": "LAI CHING-TE CROSS-STRAIT POSTURE, PLA PRESSURE, AND US SECURITY COMMITMENT"},
    {"name": "HONG KONG", "type": "COUNTRY", "scope": "BASIC LAW EROSION TRAJECTORY, ECONOMIC DECLINE, AND EMIGRATION ACCELERATION"},
    {"name": "PANAMA", "type": "COUNTRY", "scope": "MULINO AUTHORITY, CANAL SOVEREIGNTY PRESSURE, AND US TARIFF LEVERAGE"},
    {"name": "COSTA RICA", "type": "COUNTRY", "scope": "DEMOCRATIC RESILIENCE, FISCAL REFORM, AND MIGRATION TRANSIT BURDEN"},
    {"name": "GUATEMALA", "type": "COUNTRY", "scope": "AREVALO REFORM AGENDA VS ENTRENCHED ELITE, ANTI-CORRUPTION TRAJECTORY"},
    {"name": "HONDURAS", "type": "COUNTRY", "scope": "CASTRO AUTHORITY, GANG TERRITORIAL CONTROL, AND MIGRATION DRIVERS"},
    {"name": "EL SALVADOR", "type": "COUNTRY", "scope": "BUKELE BITCOIN EXPERIMENT, PRISON STATE SUSTAINABILITY, AND GANG ELIMINATION"},
    {"name": "BELIZE", "type": "COUNTRY", "scope": "BRICENO AUTHORITY, GUATEMALA TERRITORIAL CLAIM, AND REEF PRESERVATION"},
    {"name": "JAMAICA", "type": "COUNTRY", "scope": "HOLNESS AUTHORITY, GANG VIOLENCE, AND REPARATIONS DIPLOMACY"},
    {"name": "TRINIDAD", "type": "COUNTRY", "scope": "ROWLEY AUTHORITY, LNG SECTOR DECLINE, AND GANG VIOLENCE TRAJECTORY"},
    {"name": "BAHAMAS", "type": "COUNTRY", "scope": "DAVIS AUTHORITY, HURRICANE VULNERABILITY, AND OFFSHORE FINANCE REGULATION"},
    {"name": "BARBADOS", "type": "COUNTRY", "scope": "MOTTLEY AUTHORITY, REPARATIONS LEADERSHIP, AND CLIMATE VULNERABILITY"},
    {"name": "GUYANA", "type": "COUNTRY", "scope": "ALI AUTHORITY, OIL BOOM GOVERNANCE CAPACITY, AND VENEZUELA TERRITORIAL CLAIM"},
    {"name": "SURINAME", "type": "COUNTRY", "scope": "SANTOKHI AUTHORITY, OIL DISCOVERY GOVERNANCE, AND DUTCH DEPENDENCY LEGACY"},
    {"name": "URUGUAY", "type": "COUNTRY", "scope": "ORSI GOVERNMENT CONSOLIDATION, DEMOCRATIC EXCEPTIONALISM, AND ARGENTINA EXPOSURE"},
    {"name": "PARAGUAY", "type": "COUNTRY", "scope": "SANTIAGO PENA AUTHORITY, BRAZIL-ARGENTINA DEPENDENCY, AND CORRUPTION ENTRENCHMENT"},
    {"name": "BOTSWANA", "type": "COUNTRY", "scope": "BOKO GOVERNMENT CONSOLIDATION, DIAMOND DEPENDENCY DIVERSIFICATION, AND REGIONAL STABILITY"},
    {"name": "NAMIBIA", "type": "COUNTRY", "scope": "NETUMBO NANDI-NDAITWAH HISTORIC ELECTION, LAND REFORM, AND GERMAN REPARATIONS"},
    {"name": "RWANDA", "type": "COUNTRY", "scope": "KAGAME AUTHORITY, DRC PROXY WAR, AND TECH HUB AMBITION"},
    {"name": "UGANDA", "type": "COUNTRY", "scope": "MUSEVENI SUCCESSION UNCERTAINTY, OIL PIPELINE CONTROVERSY, AND LGBTQ+ LAW ISOLATION"},
    {"name": "BURUNDI", "type": "COUNTRY", "scope": "NDAYISHIMIYE AUTHORITY, GREAT LAKES STABILITY ROLE, AND HUMANITARIAN FRAGILITY"},
    {"name": "MALAWI", "type": "COUNTRY", "scope": "CHAKWERA AUTHORITY, CHRONIC FOOD INSECURITY, AND DONOR DEPENDENCY"},
    {"name": "MADAGASCAR", "type": "COUNTRY", "scope": "RAJOELINA AUTHORITY, ELECTION LEGITIMACY, AND VANILLA SECTOR DEPENDENCY"},
    {"name": "MAURITIUS", "type": "COUNTRY", "scope": "JUGNAUTH AUTHORITY, OFFSHORE FINANCIAL CENTRE REFORM, AND CHAGOS SOVEREIGNTY"},
    {"name": "SEYCHELLES", "type": "COUNTRY", "scope": "RAMKALAWAN AUTHORITY, CLIMATE VULNERABILITY, AND OCEAN ECONOMY"},
    {"name": "CAPE VERDE", "type": "COUNTRY", "scope": "NEVES AUTHORITY, REMITTANCE DEPENDENCY, AND ATLANTIC TRANSIT HUB"},
    {"name": "SAO TOME", "type": "COUNTRY", "scope": "TROVOADA AUTHORITY, OIL EXPECTATION COLLAPSE, AND LUSOPHONE BRIDGE ROLE"},
    {"name": "EQUATORIAL GUINEA", "type": "COUNTRY", "scope": "OBIANG DYNASTIC SUCCESSION, OIL REVENUE DECLINE, AND ISOLATION"},
    {"name": "GABON", "type": "COUNTRY", "scope": "JUNTA CONSOLIDATION POST-COUP, FOREST RESOURCE LEVERAGE, AND FRENCH WITHDRAWAL"},
    {"name": "COMOROS", "type": "COUNTRY", "scope": "AZALI AUTHORITY, MAYOTTE MIGRATION PRESSURE, AND ARAB LEAGUE DEPENDENCY"},
    {"name": "DJIBOUTI", "type": "COUNTRY", "scope": "GUELLEH AUTHORITY, STRATEGIC BASE COMPETITION, AND PORT LEVERAGE"},
    {"name": "ERITREA", "type": "COUNTRY", "scope": "AFWERKI ISOLATION, TIGRAY WAR AFTERMATH, AND MILITARY CONSCRIPTION EXODUS"},
    {"name": "CENTRAL AFRICAN REPUBLIC", "type": "COUNTRY", "scope": "TOUADERA WAGNER DEPENDENCY, TERRITORIAL CONTROL PATCHWORK, AND RESOURCE EXTRACTION"},
    {"name": "CHAD", "type": "COUNTRY", "scope": "MAHAMAT DEBY CONSOLIDATION, SAHEL INSTABILITY ANCHOR, AND FRENCH WITHDRAWAL"},
    {"name": "BURKINA FASO", "type": "COUNTRY", "scope": "TRAORE JUNTA, AES ALLIANCE WITH MALI AND NIGER, AND JIHADIST TERRITORIAL CONTROL"},
    {"name": "GUINEA", "type": "COUNTRY", "scope": "MAMADI DOUMBOUYA CONSOLIDATION, TRANSITION TIMELINE PRESSURE, AND BAUXITE LEVERAGE"},
    {"name": "GUINEA-BISSAU", "type": "COUNTRY", "scope": "EMBALO AUTHORITY, DRUG TRANSIT HUB STATUS, AND ECOWAS RELATIONS"},
    {"name": "SIERRA LEONE", "type": "COUNTRY", "scope": "BIO AUTHORITY, POST-ELECTION TENSION, AND IMF DEPENDENCY"},
    {"name": "LIBERIA", "type": "COUNTRY", "scope": "BOAKAI AUTHORITY, RUBBER AND TIMBER DEPENDENCY, AND DEMOCRATIC CONSOLIDATION"},
    {"name": "GAMBIA", "type": "COUNTRY", "scope": "BARROW AUTHORITY, JANJAWEED TRIAL LEGACY, AND SENEGAL DEPENDENCY"},
    {"name": "TOGO", "type": "COUNTRY", "scope": "GNASSINGBE DYNASTY SURVIVAL, CONSTITUTIONAL CHANGE, AND SAHEL BUFFER ROLE"},
    {"name": "BENIN", "type": "COUNTRY", "scope": "TALON AUTHORITY, JIHADIST NORTHERN INCURSION, AND PORT DEPENDENCY"},
    {"name": "IVORY COAST", "type": "COUNTRY", "scope": "OUATTARA SUCCESSION PLANNING, ECONOMIC LEADERSHIP, AND COCOA PRICE LEVERAGE"},
    {"name": "MAURITANIA", "type": "COUNTRY", "scope": "GHAZOUANI AUTHORITY, SAHEL STABILITY ROLE, AND GAS DISCOVERY TRAJECTORY"},
    {"name": "WESTERN SAHARA", "type": "COUNTRY", "scope": "POLISARIO RECOGNITION EROSION, MOROCCAN ADMINISTRATIVE CONTROL, AND UN PROCESS"},
    {"name": "SOUTH SUDAN", "type": "COUNTRY", "scope": "KIIR-MACHAR POWER SHARING FRAGILITY, OIL DEPENDENCY, AND HUMANITARIAN COLLAPSE"},
    {"name": "ESWATINI", "type": "COUNTRY", "scope": "MSWATI III ABSOLUTE MONARCHY, PRO-DEMOCRACY MOVEMENT, AND SOUTH AFRICA DEPENDENCY"},
    {"name": "LESOTHO", "type": "COUNTRY", "scope": "COALITION FRAGILITY, SOUTH AFRICA ENCLAVE DEPENDENCY, AND WATER LEVERAGE"},
    {"name": "SOLOMON ISLANDS", "type": "COUNTRY", "scope": "SOGAVARE CHINA PACT LEGACY, AUSTRALIAN PRESSURE, AND POLITICAL STABILITY"},
    {"name": "VANUATU", "type": "COUNTRY", "scope": "COALITION DYSFUNCTION RECORD, CHINA INFLUENCE, AND CLIMATE VULNERABILITY"},
    {"name": "SAMOA", "type": "COUNTRY", "scope": "FIAME AUTHORITY, CHINA INFRASTRUCTURE DEPENDENCY, AND PACIFIC UNITY ROLE"},
    {"name": "TONGA", "type": "COUNTRY", "scope": "TUPOU VI MONARCHY, CHINA DEBT, AND TSUNAMI RECOVERY"},
    {"name": "KIRIBATI", "type": "COUNTRY", "scope": "MAAMAU AUTHORITY, CLIMATE EXISTENTIAL CRISIS, AND CHINA ALIGNMENT"},
    {"name": "MICRONESIA", "type": "COUNTRY", "scope": "US COMPACT DEPENDENCY, CHINA OUTREACH RESISTANCE, AND CLIMATE VULNERABILITY"},
    {"name": "PALAU", "type": "COUNTRY", "scope": "WHIPPS AUTHORITY, US MILITARY PRESENCE, AND TAIWAN RECOGNITION"},
    {"name": "MARSHALL ISLANDS", "type": "COUNTRY", "scope": "CABINET AUTHORITY, NUCLEAR TEST LEGACY, AND COMPACT RENEGOTIATION"},
    {"name": "NAURU", "type": "COUNTRY", "scope": "ADEANG AUTHORITY, TAIWAN RECOGNITION, AND PHOSPHATE DEPLETION LEGACY"},
    {"name": "TUVALU", "type": "COUNTRY", "scope": "FELETI AUTHORITY, AUSTRALIA TREATY, AND CLIMATE ERASURE TIMELINE"},
    {"name": "TRUMP, D.", "type": "PRESIDENT", "scope": "US POLITICAL SURVIVAL THROUGH 2026 MIDTERMS"},
    {"name": "PUTIN, V.", "type": "PRESIDENT", "scope": "KREMLIN POWER RETENTION 36-MONTH PROJECTION"},
    {"name": "XI JINPING", "type": "PRESIDENT", "scope": "CCP CONSOLIDATION AND TAIWAN POLICY TRAJECTORY"},
    {"name": "MACRON, E.", "type": "PRESIDENT", "scope": "FIFTH REPUBLIC AUTHORITY THROUGH 2027"},
    {"name": "MODI, N.", "type": "PRIME MINISTER", "scope": "BJP COALITION DURABILITY AND REGIONAL STABILITY"},
    {"name": "ERDOGAN, R.", "type": "PRESIDENT", "scope": "AKP CONSOLIDATION AND ECONOMIC STABILISATION"},
    {"name": "ZELENSKY, V.", "type": "PRESIDENT", "scope": "WAR LEADERSHIP LEGITIMACY AND WESTERN SUPPORT"},
    {"name": "STARMER, K.", "type": "PRIME MINISTER", "scope": "LABOUR GOVERNMENT REFORM CAPACITY AND APPROVAL"},
    {"name": "NETANYAHU, B.", "type": "PRIME MINISTER", "scope": "WAR CABINET SURVIVAL AND JUDICIAL THREAT VECTOR"},
    {"name": "SCHOLZ, O.", "type": "CHANCELLOR", "scope": "COALITION COLLAPSE AFTERMATH AND ELECTION RECOVERY"},
    {"name": "MELONI, G.", "type": "PRIME MINISTER", "scope": "FAR-RIGHT GOVERNANCE CONSOLIDATION IN EU"},
    {"name": "ORBAN, V.", "type": "PRIME MINISTER", "scope": "EU LEVERAGE AND ILLIBERAL CONSOLIDATION"},
    {"name": "FARAGE, N.", "type": "OPPOSITION", "scope": "REFORM UK ELECTORAL TRAJECTORY THROUGH 2029"},
    {"name": "LE PEN, M.", "type": "OPPOSITION", "scope": "RN PATH TO ELYSEE THROUGH 2027"},
    {"name": "LULA, L.", "type": "PRESIDENT", "scope": "PT COALITION DURABILITY AND AMAZON GOVERNANCE"},
    {"name": "MILEI, J.", "type": "PRESIDENT", "scope": "ANARCHO-CAPITALIST SHOCK THERAPY SURVIVAL"},
    {"name": "MBS", "type": "CROWN PRINCE", "scope": "VISION 2030 DELIVERY AND SUCCESSION CONSOLIDATION"},
    {"name": "KHAMENEI, A.", "type": "SUPREME LEADER", "scope": "CLERICAL REGIME LEGITIMACY AND NUCLEAR LEVERAGE"},
    {"name": "NATO EASTERN FLANK", "type": "GEO_NODE", "scope": "DETERRENCE CREDIBILITY AND ARTICLE 5 COHESION"},
    {"name": "TAIWAN STRAIT", "type": "GEO_NODE", "scope": "ESCALATION TRAJECTORY AND US COMMITMENT CREDIBILITY"},
    {"name": "UK DOMESTIC FRACTURE", "type": "GEO_NODE", "scope": "DEVOLUTION PRESSURE AND POPULIST INSURGENCY"},
    {"name": "EU FRAGMENTATION", "type": "GEO_NODE", "scope": "INSTITUTIONAL COHESION AND FAR-RIGHT CAPTURE RISK"},
    {"name": "BRICS EXPANSION", "type": "GEO_NODE", "scope": "ALTERNATIVE ORDER CONSOLIDATION VS WESTERN SYSTEM"},
    {"name": "RED SEA CORRIDOR", "type": "GEO_NODE", "scope": "HOUTHI DISRUPTION CAPACITY AND SHIPPING RISK"},
    {"name": "ARCTIC SOVEREIGNTY", "type": "GEO_NODE", "scope": "GREAT POWER COMPETITION AND RESOURCE CLAIM CONFLICTS"},
    {"name": "SAHEL COLLAPSE", "type": "GEO_NODE", "scope": "JUNTA CONSOLIDATION AND JIHADIST TERRITORIAL EXPANSION"},
    {"name": "MUSK, E.", "type": "FIGURE", "scope": "POLITICAL INFLUENCE REACH AND DOGE INSTITUTIONAL IMPACT"},
    {"name": "SOROS, G.", "type": "FIGURE", "scope": "OPEN SOCIETY NETWORK AND FAR-RIGHT COUNTER-PRESSURE"},
    {"name": "SCHWAB, K.", "type": "FIGURE", "scope": "WEF RELEVANCE AND DAVOS CONSENSUS DURABILITY"},
    {"name": "THIEL, P.", "type": "FIGURE", "scope": "TECH-RIGHT NETWORK INFLUENCE AND NRX IDEOLOGICAL SPREAD"},
    {"name": "PRIGOZHIN LEGACY", "type": "FIGURE", "scope": "WAGNER SUCCESSOR STRUCTURES AND KREMLIN CONTROL"},
]

# ─────────────────────────────────────────────
# REFINEMENT MEMORY — loads past scores to inform prompting
# ─────────────────────────────────────────────

# load_refinement_context() replaced by Algorithm 1 above

# ─────────────────────────────────────────────
# GAP REPORT GENERATION
# ─────────────────────────────────────────────

def generate_gap_report(subject: dict, refinement_context: str) -> dict:
    """Call Claude API to produce structured GAP report data."""
    client = anthropic.Anthropic(api_key=os.getenv("ANTHROPIC_API_KEY", ""))

    now = datetime.datetime.utcnow()
    ref_id = f"ICE-{subject['type'][:3].upper()}-{now.strftime('%Y%m%d%H%M')}"
    title  = f"//INTERSTITIAL \"{subject['name']}\" {now.strftime('%H:%M %d/%m/%Y')}"

    refinement_note = ""
    ctx = refinement_context if isinstance(refinement_context, dict) else {}
    ctx_str = ctx.get("context_str", "")
    if ctx_str:
        reset_note = ""
        if ctx.get("reset_triggered"):
            reset_note = f"\nCONTEXT RESET FIRED: recalibrate {subject['name']} from first principles. Ignore prior scores for this subject."
        anchor = ctx.get("anchor_strength", 0)
        refinement_note = f"""
SELF-REFINEMENT CONTEXT (Algorithm 1 — decay-weighted):
{ctx_str}{reset_note}
Anchor strength for this subject: {anchor:.0%}. 
If anchor >60% apply extra variance — do not simply reproduce prior scores.
"""

    system_prompt = """You are the INTERSTITIAL CYCLE ENGINE, a quantified political pressure analysis system.
You produce structured JSON node assessments for political and geopolitical cycle reports.
Be analytically rigorous, terse, direct. No hedging. No filler. Pure structured intelligence output.
Return ONLY valid JSON. No preamble. No explanation outside the JSON."""

    user_prompt = f"""Produce a full GAP Report for the following subject as structured JSON.

SUBJECT: {subject['name']}
TYPE: {subject['type']}
SCOPE: {subject['scope']}
REF: {ref_id}
DATE: {now.strftime('%Y-%m-%d')}
{refinement_note}

Define 6 nodes appropriate to this subject. Each node must be analytically distinct and 
structurally meaningful for this specific subject. Choose node names that reflect the 
REAL structural forces governing this subject's trajectory.

Return this exact JSON schema:

{{
  "subject": "{subject['name']}",
  "type": "{subject['type']}",
  "scope": "{subject['scope']}",
  "ref": "{ref_id}",
  "title": "{title}",
  "timestamp": "{now.isoformat()}",
  "nodes": [
    {{
      "id": "N-01",
      "name": "SHORT NODE NAME IN CAPS",
      "status": "ONE OF: STRONG/DECAYING/FRAGILE/CONTESTED/CRITICAL PATH/CONDITIONAL/ERODED/EXPANDING/TIGHTENING/LATENT/RECOVERING/WEAKENED/ELEVATED/EMBRYONIC/ACCELERANT/ACTIVE/TERMINAL/CRITICAL",
      "score": 0,
      "assessment": "2-3 sentence factual assessment. Terse. Evidence-based. No hedging.",
      "gap": "1-2 sentence gap analysis. What is missing. What threatens this node."
    }}
  ],
  "composite_score": 0,
  "scenarios": [
    {{"id": "A", "label": "SCENARIO LABEL IN CAPS", "probability": 0, "outcome": "SURVIVAL/HIGH/MEDIUM/LOW/CRITICAL/TERMINAL"}},
    {{"id": "B", "label": "SCENARIO LABEL IN CAPS", "probability": 0, "outcome": "MEDIUM"}},
    {{"id": "C", "label": "SCENARIO LABEL IN CAPS", "probability": 0, "outcome": "LOW"}},
    {{"id": "D", "label": "SCENARIO LABEL IN CAPS", "probability": 0, "outcome": "CRITICAL"}}
  ],
  "master_node": "N-0X",
  "synthesis": "3-4 sentence synthesis. Name the master node. State the binding constraint. State the optimal window or failure trigger.",
  "nodes_edge_weights": {{
    "N-01": 0.17, "N-02": 0.17, "N-03": 0.17,
    "N-04": 0.17, "N-05": 0.17, "N-06": 0.15
  }}
}}

Score rules: All node scores 15-95. Composite = weighted mean of node scores. 
Scenario probabilities must sum to 100 or close. Be analytically honest. 
Avoid clustering all scores near 50 — differentiate meaningfully."""

    response = client.messages.create(
        model="claude-opus-4-5",
        max_tokens=3000,
        system=system_prompt,
        messages=[{"role": "user", "content": user_prompt}]
    )

    raw = response.content[0].text.strip()
    # Strip markdown fences if present
    raw = re.sub(r'^```json\s*', '', raw)
    raw = re.sub(r'^```\s*', '', raw)
    raw = re.sub(r'\s*```$', '', raw)

    data = json.loads(raw)
    return data

# ─────────────────────────────────────────────
# RENDER: BRAND-COMPLIANT PAGE IMAGES & MARKOV GRAPH
# Interstitial brand spec v1.0 — Liberation Sans, #0A0A0A, 2400px
# ─────────────────────────────────────────────


import os, textwrap
from pathlib import Path

import numpy as np
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
from matplotlib.image import imread as mpl_imread
from PIL import Image, ImageDraw, ImageFont

# ── Paths ──────────────────────────────────────────────────────────────────────
_DIR = Path(__file__).parent

SPIRAL_PATH   = _DIR / "spiral_symbol.png"
WORDMARK_PATH = _DIR / "logo_dark_wordmark.png"

FONT_BOLD = "/usr/share/fonts/truetype/liberation/LiberationSans-Bold.ttf"
FONT_REG  = "/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf"

# ── Palette ────────────────────────────────────────────────────────────────────
BG     = (10,  10,  10)
WHITE  = (255, 255, 255)
G70    = (77,  77,  77)
G50    = (128, 128, 128)
G15    = (217, 217, 217)
TRACK  = (28,  28,  28)

# ── Font cache ─────────────────────────────────────────────────────────────────
_FC = {}
def F(size, bold=False):
    k = (size, bold)
    if k not in _FC:
        p = FONT_BOLD if bold else FONT_REG
        try:
            _FC[k] = ImageFont.truetype(p, size)
        except Exception:
            _FC[k] = ImageFont.load_default()
    return _FC[k]

# ── Score → grey ───────────────────────────────────────────────────────────────
def sg(pct):
    """Score → RGB grey tuple, 0%=dim, 100%=white"""
    v = int(55 + min(pct, 100) / 100 * 200)
    return (v, v, v)

def sg_f(pct):
    v = (55 + min(pct, 100) / 100 * 200) / 255
    return (v, v, v)

# ── Layout constants ───────────────────────────────────────────────────────────
W   = 2400
PAD = 84
CW  = W - PAD * 2   # 2232
HH  = 160           # header height

# ── Header ─────────────────────────────────────────────────────────────────────
def draw_header(img, draw, page_ref=""):
    """Spiral icon + Interstitial wordmark + rule. No grey box, no badge."""
    icon_sz = 108
    icon_y  = (HH - icon_sz) // 2

    # Spiral symbol
    if SPIRAL_PATH.exists():
        sp = Image.open(str(SPIRAL_PATH)).convert("RGB").resize(
            (icon_sz, icon_sz), Image.LANCZOS)
        img.paste(sp, (PAD, icon_y))
    else:
        draw.rectangle([PAD, icon_y, PAD+icon_sz, icon_y+icon_sz], outline=WHITE)

    tx = PAD + icon_sz + 28

    # "Interstitial" — bold 34pt
    draw.text((tx, icon_y + 4), "Interstitial",
              font=F(50, bold=True), fill=WHITE)

    # "MEASURING THE FUTURE" — regular 11pt, tracked feel via spacing
    draw.text((tx, icon_y + 68), "MEASURING THE FUTURE",
              font=F(16), fill=G50)

    # Page ref — micro grey, top right
    if page_ref:
        rw = draw.textlength(page_ref, font=F(11))
        draw.text((W - PAD - rw, 24), page_ref, font=F(16), fill=G50)

    # White rule under header — 1px
    draw.line([(PAD, HH - 1), (W - PAD, HH - 1)], fill=WHITE, width=1)

# ── Score bar ──────────────────────────────────────────────────────────────────
def draw_bar(draw, x, y, w, pct, h=24):
    """Full-width track, solid fill scaled to pct, % label right of fill."""
    draw.rectangle([x, y, x + w, y + h], fill=TRACK)
    fw = int(w * pct / 100)
    if fw > 0:
        draw.rectangle([x, y, x + fw, y + h], fill=sg(pct))
    # % label
    lbl = f"{pct}%"
    lx  = x + fw + 10 if fw < w - 60 else x + fw - 50
    ly  = y + h // 2 - 8
    draw.text((lx, ly), lbl, font=F(21, bold=True), fill=WHITE)

# ── Thin rule ──────────────────────────────────────────────────────────────────
def rule(draw, y, col=G70, weight=1):
    draw.line([(PAD, y), (W - PAD, y)], fill=col, width=weight)

# ── Text wrap ─────────────────────────────────────────────────────────────────
def wrap(text, chars=108):
    return textwrap.wrap(str(text), width=chars)

# ── Pre-calculate page height ──────────────────────────────────────────────────
def calc_height(blocks):
    h = HH + 40
    for b in blocks:
        t = b[0]
        if   t == 'space':   h += b[1]
        elif t == 'rule':    h += 30
        elif t == 'h1':      h += 68
        elif t == 'h2':      h += 48
        elif t == 'meta':    h += 34
        elif t == 'kv':      h += 44
        elif t == 'body':    h += len(wrap(b[1])) * 34 + 12
        elif t == 'node':
            nd = b[1]
            h += 42                                    # id + name row
            h += 40                                    # bar
            h += len(wrap(nd.get('assessment',''))) * 34 + 12
            h += len(wrap('GAP  ' + nd.get('gap',''))) * 34 + 32
        elif t == 'summary':
            h += len(b[1]) * 58 + 54                  # per-node 2 rows
        elif t == 'scenarios':
            h += len(b[1]) * 76 + 24
        elif t == 'synthesis':
            h += len(wrap(b[1], 100)) * 40 + 48
    return max(h + 60, 1000)

# ── Page render ────────────────────────────────────────────────────────────────
def render_page(blocks, page_num, total, out_path):
    H   = calc_height(blocks)
    img = Image.new('RGB', (W, H), BG)
    drw = ImageDraw.Draw(img)

    draw_header(img, drw, page_ref=f"{page_num} / {total}")

    y = HH + 54

    for b in blocks:
        t = b[0]

        # ── space ─────────────────────────────────────────────────────────────
        if t == 'space':
            y += b[1]

        # ── rule ──────────────────────────────────────────────────────────────
        elif t == 'rule':
            col, wt = (b[1], b[2]) if len(b) > 2 else (G70, 1)
            rule(drw, y + 12, col, wt)
            y += 30

        # ── h1: section heading — white bold 18pt ─────────────────────────────
        elif t == 'h1':
            drw.text((PAD, y), b[1].upper(), font=F(27, bold=True), fill=WHITE)
            y += 68

        # ── h2: sub-heading — grey15 bold 14pt ───────────────────────────────
        elif t == 'h2':
            drw.text((PAD, y), b[1].upper(), font=F(21, bold=True), fill=G15)
            y += 48

        # ── meta: small grey line ─────────────────────────────────────────────
        elif t == 'meta':
            drw.text((PAD, y), b[1], font=F(17), fill=G50)
            y += 34

        # ── kv: key  value pair ───────────────────────────────────────────────
        elif t == 'kv':
            key, val = b[1], b[2]
            kw = drw.textlength(key.upper() + "  ", font=F(19, bold=True))
            drw.text((PAD, y), key.upper() + "  ", font=F(19, bold=True), fill=G50)
            drw.text((PAD + kw, y), val.upper(), font=F(19, bold=False), fill=WHITE)
            y += 44

        # ── body: grey70 regular 13pt wrapped ────────────────────────────────
        elif t == 'body':
            for ln in wrap(b[1]):
                drw.text((PAD, y), ln, font=F(20), fill=G70)
                y += 34
            y += 12

        # ── node: full node block ─────────────────────────────────────────────
        elif t == 'node':
            nd     = b[1]
            nid    = nd['id']
            nname  = nd['name']
            status = nd.get('status', '').upper()
            score  = nd.get('score', 0)
            assess = nd.get('assessment', '')
            gap    = nd.get('gap', '')

            # Row 1: N-01  NODE NAME                            STATUS
            id_str = nid + "  "
            iw     = drw.textlength(id_str, font=F(22, bold=True))
            drw.text((PAD, y), id_str, font=F(22, bold=True), fill=WHITE)
            drw.text((PAD + iw, y), nname.upper(), font=F(22), fill=G15)
            sw = drw.textlength(status, font=F(16))
            drw.text((W - PAD - sw, y + 4), status, font=F(16), fill=G50)
            y += 42

            # Row 2: score bar — full CW
            draw_bar(drw, PAD, y, CW, score, h=24)
            y += 40

            # Assessment
            for ln in wrap(assess):
                drw.text((PAD, y), ln, font=F(20), fill=G70)
                y += 34
            y += 12

            # GAP
            gap_label = "GAP  "
            glw = drw.textlength(gap_label, font=F(18, bold=True))
            gap_lines = wrap(gap, chars=100)
            for i, ln in enumerate(gap_lines):
                if i == 0:
                    drw.text((PAD, y), gap_label, font=F(18, bold=True), fill=G50)
                    drw.text((PAD + glw, y), ln, font=F(18), fill=G50)
                else:
                    drw.text((PAD + glw, y), ln, font=F(18), fill=G50)
                y += 34
            y += 32

        # ── summary: node strength table ──────────────────────────────────────
        elif t == 'summary':
            nodes, composite = b[1], b[2]
            for nd in nodes:
                nid   = nd['id'] + "  "
                nname = nd['name'].upper()
                score = nd.get('score', 0)
                stat  = nd.get('status', '').upper()

                iw = drw.textlength(nid, font=F(19, bold=True))
                drw.text((PAD, y), nid, font=F(19, bold=True), fill=WHITE)
                drw.text((PAD + iw, y), nname, font=F(19), fill=G15)
                sw = drw.textlength(stat, font=F(15))
                drw.text((W - PAD - sw, y + 2), stat, font=F(15), fill=G50)
                y += 30

                draw_bar(drw, PAD, y, CW, score, h=18)
                y += 28

            # Composite
            rule(drw, y + 6, G70)
            y += 20
            comp_label = "COMPOSITE SCORE  "
            clw = drw.textlength(comp_label, font=F(24, bold=True))
            drw.text((PAD, y), comp_label, font=F(24, bold=True), fill=G50)
            drw.text((PAD + clw, y), f"{composite}%", font=F(24, bold=True), fill=WHITE)
            y += 54

        # ── scenarios: probability bar matrix ────────────────────────────────
        elif t == 'scenarios':
            for sc in b[1]:
                sid     = f"[{sc['id']}]"
                label   = sc.get('label', '').upper()
                prob    = sc.get('probability', 0)
                outcome = sc.get('outcome', '').upper()

                # Label row
                prefix = sid + "  "
                pw = drw.textlength(prefix, font=F(21, bold=True))
                drw.text((PAD, y), prefix, font=F(21, bold=True), fill=WHITE)
                drw.text((PAD + pw, y), label, font=F(21), fill=G15)
                ow = drw.textlength(outcome, font=F(16))
                drw.text((W - PAD - ow, y + 3), outcome, font=F(16), fill=G50)
                y += 36

                # Probability bar
                draw_bar(drw, PAD, y, CW, prob, h=22)
                y += 38

        # ── synthesis: grey15 regular 14pt ───────────────────────────────────
        elif t == 'synthesis':
            for ln in wrap(b[1], 100):
                drw.text((PAD, y), ln, font=F(21), fill=G15)
                y += 40
            y += 24

    img.save(str(out_path), "JPEG", quality=98, subsampling=0)


# ── Build page block lists ─────────────────────────────────────────────────────
def build_pages(data):
    subj      = data['subject']
    scope     = data['scope']
    ref       = data['ref']
    ts        = data['timestamp'][:10]
    nodes     = data['nodes']
    scenarios = data['scenarios']
    composite = data['composite_score']
    synthesis = data['synthesis']
    master    = data['master_node']
    n         = data.get('cycle_number', 0)
    cyc       = f"INTER//CYCLE {str(n).zfill(6)}"

    meta = f"{ts}    {ref}    {cyc}"

    def header_block(pg):
        return [
            ('meta',  meta),
            ('space', 10),
            ('kv',    'Subject', subj),
            ('kv',    'Scope',   scope),
            ('space', 12),
            ('rule',  WHITE, 1),
        ]

    # Page 1 — nodes N-01/02/03
    p1 = header_block(1) + [
        ('h1',   'Node Assessments'),
        ('rule', G70, 1),
        ('space', 10),
    ]
    for nd in nodes[:3]:
        p1 += [('node', nd), ('rule', G70, 1), ('space', 6)]

    # Page 2 — nodes N-04/05/06
    p2 = header_block(2) + [
        ('h1',   'Node Assessments  —  continued'),
        ('rule', G70, 1),
        ('space', 10),
    ]
    for nd in nodes[3:]:
        p2 += [('node', nd), ('rule', G70, 1), ('space', 6)]

    # Page 3 — summary / scenarios / synthesis
    p3 = header_block(3) + [
        ('h1',   'Node Strength Summary'),
        ('rule', G70, 1),
        ('space', 10),
        ('summary',   nodes, composite),
        ('space', 10),
        ('rule',  WHITE, 1),
        ('h1',   'Scenario Matrix  —  36-Month Horizon'),
        ('rule', G70, 1),
        ('space', 10),
        ('scenarios', scenarios),
        ('space', 10),
        ('rule',  WHITE, 1),
        ('h1',   'Synthesis'),
        ('rule', G70, 1),
        ('space', 12),
        ('synthesis', synthesis),
        ('meta', f"Master node: {master}    {ref}"),
    ]

    return [p1, p2, p3]


# ── Entry point ────────────────────────────────────────────────────────────────
def render_report_images(data, run_dir):
    run_dir = Path(run_dir)
    pages   = build_pages(data)
    paths   = []
    for i, pg in enumerate(pages, 1):
        out = run_dir / f"page_{i}.jpg"
        render_page(pg, i, 3, out)
        paths.append(out)
    return paths


# ── Markov graph ───────────────────────────────────────────────────────────────
def render_markov_graph(data, run_dir):
    """
    Brand-compliant node graph.
    - Pure #0A0A0A background
    - Liberation Sans throughout (no monospace)
    - Spiral + wordmark header
    - White solid bars on nodes, brightness-scaled borders
    - Double border on COMPOSITE
    - White arrows node→composite, dark-grey dashed inter-node
    - No footer
    """
    run_dir = Path(run_dir)

    nodes     = data['nodes']
    subject   = data['subject']
    ref       = data['ref']
    ts        = data['timestamp'][:10]
    composite = data['composite_score']
    n_val     = data.get('cycle_number', 0)
    cyc       = f"INTER//CYCLE {str(n_val).zfill(6)}"

    plt.rcParams.update({
        'font.family':   'Liberation Sans',
        'text.color':    'white',
        'figure.facecolor': '#0A0A0A',
    })

    FIG_W, FIG_H = 24, 19
    fig = plt.figure(figsize=(FIG_W, FIG_H))
    fig.patch.set_facecolor('#0A0A0A')

    # ── Header bar: spiral + wordmark ─────────────────────────────────────────
    HEADER_FRAC = 0.072
    SUB_FRAC    = 0.044
    GRAPH_BOT   = 0.055
    GRAPH_TOP   = 1.0 - HEADER_FRAC - SUB_FRAC

    if SPIRAL_PATH.exists():
        sp_ax = fig.add_axes([0.025, 1.0 - HEADER_FRAC + 0.005,
                              0.046, HEADER_FRAC - 0.010])
        sp_ax.imshow(np.array(Image.open(str(SPIRAL_PATH)).convert('RGB')),
                     aspect='auto', interpolation='lanczos')
        sp_ax.axis('off')

    fig.text(0.083, 1.0 - HEADER_FRAC * 0.36, "Interstitial",
             fontsize=19, fontweight='bold', color='white',
             fontfamily='Liberation Sans', va='center')
    fig.text(0.083, 1.0 - HEADER_FRAC * 0.76, "MEASURING THE FUTURE",
             fontsize=7.5, color='#808080',
             fontfamily='Liberation Sans', va='center')

    # White rule under header
    rule_y = 1.0 - HEADER_FRAC
    fig.add_artist(plt.Line2D(
        [0.025, 0.975], [rule_y, rule_y],
        transform=fig.transFigure,
        color='white', linewidth=0.7, zorder=10))

    # Subject + cycle metadata — strip directly below rule
    sub_mid  = rule_y - SUB_FRAC * 0.32
    meta_mid = rule_y - SUB_FRAC * 0.72
    fig.text(0.50, sub_mid, subject.upper(),
             fontsize=12, fontweight='bold', color='white',
             ha='center', fontfamily='Liberation Sans')
    fig.text(0.50, meta_mid, f"{ref}    {ts}    {cyc}",
             fontsize=7.5, color='#808080',
             ha='center', fontfamily='Liberation Sans')

    # Dim rule separating header from graph
    fig.add_artist(plt.Line2D(
        [0.025, 0.975], [GRAPH_TOP, GRAPH_TOP],
        transform=fig.transFigure,
        color='#2A2A2A', linewidth=0.5, zorder=10))

    # ── Graph axes ─────────────────────────────────────────────────────────────
    ax = fig.add_axes([0.025, GRAPH_BOT, 0.95, GRAPH_TOP - GRAPH_BOT])
    ax.set_facecolor('#0A0A0A')
    ax.set_xlim(0, 10)
    ax.set_ylim(0, 10)
    ax.axis('off')

    # Node ring layout
    cx, cy   = 5.0, 5.3
    radius   = 3.15
    angles   = [90, 30, 330, 270, 210, 150]
    NODE_W   = 1.80
    NODE_H   = 1.20

    positions = {}
    for i, nd in enumerate(nodes):
        a = np.radians(angles[i])
        positions[nd['id']] = (cx + radius * np.cos(a),
                               cy + radius * np.sin(a))
    positions['COMPOSITE'] = (cx, cy)

    # ── Draw node box ──────────────────────────────────────────────────────────
    def draw_node(key, nd_data, is_comp=False):
        x, y   = positions[key]
        pct    = nd_data.get('score', composite)
        name   = nd_data.get('name', key)
        gc     = sg_f(pct)
        lw     = 2.0 if is_comp else 1.2

        if is_comp:
            # Outer border (double border effect)
            ax.add_patch(mpatches.FancyBboxPatch(
                (x - NODE_W/2 - 0.14, y - NODE_H/2 - 0.14),
                NODE_W + 0.28, NODE_H + 0.28,
                boxstyle="square,pad=0", linewidth=0.9,
                edgecolor=(0.40, 0.40, 0.40), facecolor='#0A0A0A'))

        ax.add_patch(mpatches.FancyBboxPatch(
            (x - NODE_W/2, y - NODE_H/2), NODE_W, NODE_H,
            boxstyle="square,pad=0", linewidth=lw,
            edgecolor=gc, facecolor='#0A0A0A'))

        if is_comp:
            ax.text(x, y + 0.17, "COMPOSITE",
                    ha='center', va='center',
                    fontsize=11, fontweight='bold', color='white',
                    fontfamily='Liberation Sans')
            ax.text(x, y - 0.12, f"{composite}%",
                    ha='center', va='center',
                    fontsize=18, fontweight='bold', color='white',
                    fontfamily='Liberation Sans')
        else:
            # Node ID bold, name regular
            ax.text(x, y + 0.34, key,
                    ha='center', va='center',
                    fontsize=9, fontweight='bold', color='white',
                    fontfamily='Liberation Sans')
            short = name[:18] if len(name) > 18 else name
            ax.text(x, y + 0.14, short,
                    ha='center', va='center',
                    fontsize=7.5, color=gc,
                    fontfamily='Liberation Sans')

            # Score bar — solid fill, brightness-scaled
            bx      = x - NODE_W/2 + 0.08
            by      = y - NODE_H/2 + 0.09
            bw_full = NODE_W - 0.16
            bw_fill = bw_full * (pct / 100)
            ax.add_patch(mpatches.Rectangle(
                (bx, by), bw_full, 0.13, color=(0.11, 0.11, 0.11)))
            ax.add_patch(mpatches.Rectangle(
                (bx, by), bw_fill, 0.13, color=gc))
            ax.text(x, by + 0.065, f"{pct}%",
                    ha='center', va='center',
                    fontsize=7, color='white',
                    fontfamily='Liberation Sans')

    # Inter-node edges — dashed dark grey
    edge_pairs = [
        (nodes[0]['id'], nodes[1]['id']),
        (nodes[1]['id'], nodes[2]['id']),
        (nodes[2]['id'], nodes[3]['id']),
        (nodes[3]['id'], nodes[4]['id']),
        (nodes[4]['id'], nodes[5]['id']),
        (nodes[5]['id'], nodes[0]['id']),
        (nodes[0]['id'], nodes[3]['id']),
        (nodes[1]['id'], nodes[4]['id']),
    ]
    for f_id, t_id in edge_pairs:
        x0, y0 = positions[f_id]
        x1, y1 = positions[t_id]
        ax.annotate('', xy=(x1, y1), xytext=(x0, y0),
                    arrowprops=dict(
                        arrowstyle='->', color='#303030',
                        lw=0.8, linestyle='--',
                        connectionstyle='arc3,rad=0.13',
                        shrinkA=46, shrinkB=46))

    # Node → Composite edges — solid white
    weights = data.get('nodes_edge_weights', {})
    for nd in nodes:
        x0, y0 = positions[nd['id']]
        x1, y1 = positions['COMPOSITE']
        w = weights.get(nd['id'], 0.17)
        ax.annotate('', xy=(x1, y1), xytext=(x0, y0),
                    arrowprops=dict(
                        arrowstyle='->', color='white',
                        lw=1.1,
                        connectionstyle='arc3,rad=0.05',
                        shrinkA=46, shrinkB=52))
        mx, my = (x0 + x1) / 2, (y0 + y1) / 2
        ax.text(mx, my, f"{w:.2f}",
                ha='center', va='center',
                fontsize=5.5, color='#4D4D4D',
                fontfamily='Liberation Sans',
                bbox=dict(boxstyle='square,pad=0.1',
                          fc='#0A0A0A', ec='none'))

    # Draw all nodes
    for nd in nodes:
        draw_node(nd['id'], nd)
    draw_node('COMPOSITE', {'score': composite, 'name': 'COMPOSITE'}, is_comp=True)

    # ── Legend ─────────────────────────────────────────────────────────────────
    lx, ly = 0.30, 0.50
    ax.plot([lx, lx + 0.30], [ly + 0.20, ly + 0.20],
            color='white', lw=0.9)
    ax.annotate('', xy=(lx + 0.30, ly + 0.20), xytext=(lx, ly + 0.20),
                arrowprops=dict(arrowstyle='->', color='white', lw=0.9))
    ax.text(lx + 0.38, ly + 0.20, 'Node  →  Composite  (weighted)',
            fontsize=6.5, color='#808080', va='center',
            fontfamily='Liberation Sans')
    ax.plot([lx, lx + 0.30], [ly + 0.04, ly + 0.04],
            color='#303030', lw=0.8, linestyle='--')
    ax.text(lx + 0.38, ly + 0.04, 'Inter-node dependency',
            fontsize=6.5, color='#4D4D4D', va='center',
            fontfamily='Liberation Sans')

    out = run_dir / "markov_graph.jpg"
    plt.savefig(str(out), dpi=200, bbox_inches='tight',
                facecolor='#0A0A0A', edgecolor='none')
    plt.close()
    return out

# ─────────────────────────────────────────────
# SPREADSHEET LOGGING
# ─────────────────────────────────────────────

def init_spreadsheet():
    if DATA_FILE.exists():
        return
    wb = openpyxl.Workbook()

    # RUNS sheet
    ws = wb.active
    ws.title = "RUNS"
    headers = [
        "RUN_ID", "TIMESTAMP", "SUBJECT", "TYPE", "SCOPE", "REF",
        "N01_NAME", "N01_SCORE", "N01_STATUS",
        "N02_NAME", "N02_SCORE", "N02_STATUS",
        "N03_NAME", "N03_SCORE", "N03_STATUS",
        "N04_NAME", "N04_SCORE", "N04_STATUS",
        "N05_NAME", "N05_SCORE", "N05_STATUS",
        "N06_NAME", "N06_SCORE", "N06_STATUS",
        "COMPOSITE_SCORE", "MASTER_NODE", "NODE_SCORES",
        "SCENARIO_A_PROB", "SCENARIO_B_PROB", "SCENARIO_C_PROB", "SCENARIO_D_PROB",
        "TITLE", "OUTPUT_DIR"
    ]
    hdr_font = Font(bold=True, color="FFFFFF")
    hdr_fill = PatternFill("solid", start_color="1A1A1A")
    ws.append(headers)
    for cell in ws[1]:
        cell.font  = hdr_font
        cell.fill  = hdr_fill
        cell.alignment = Alignment(horizontal='center')

    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    # ANALYSIS sheet
    ws2 = wb.create_sheet("ANALYSIS")
    ws2.append(["METRIC", "VALUE", "NOTES"])
    ws2.append(["Total runs", '=COUNTA(RUNS!A:A)-1', "Excludes header"])
    ws2.append(["Avg composite score", '=AVERAGE(RUNS!Y:Y)', "All runs"])
    ws2.append(["Highest composite", '=MAX(RUNS!Y:Y)', "Best scoring subject"])
    ws2.append(["Lowest composite", '=MIN(RUNS!Y:Y)', "Weakest subject"])
    ws2.append(["Most recent run", '=MAX(RUNS!B:B)', "Latest timestamp"])

    ws2["A1"].font = Font(bold=True)
    ws2["B1"].font = Font(bold=True)
    ws2["C1"].font = Font(bold=True)

    wb.save(str(DATA_FILE))

def log_to_spreadsheet(data: dict, run_dir: Path):
    init_spreadsheet()
    wb = openpyxl.load_workbook(str(DATA_FILE))
    ws = wb["RUNS"]

    nodes = data['nodes']
    scenarios = {sc['id']: sc['probability'] for sc in data['scenarios']}
    node_score_str = ", ".join(f"{n['id']}:{n['score']}%" for n in nodes)

    run_id = f"RUN-{datetime.datetime.utcnow().strftime('%Y%m%d%H%M%S')}"
    row = [
        run_id,
        data['timestamp'],
        data['subject'],
        data['type'],
        data['scope'],
        data['ref'],
    ]
    for n in nodes:
        row += [n['name'], n['score'], n['status']]
    row += [
        data['composite_score'],
        data['master_node'],
        node_score_str,
        scenarios.get('A', ''), scenarios.get('B', ''),
        scenarios.get('C', ''), scenarios.get('D', ''),
        data['title'],
        str(run_dir),
    ]
    ws.append(row)

    # Colour composite score cell
    last_row = ws.max_row
    composite_col = 25  # Y
    cell = ws.cell(row=last_row, column=composite_col)
    score = data['composite_score']
    if score >= 70:
        cell.fill = PatternFill("solid", start_color="1A4A1A")
    elif score >= 50:
        cell.fill = PatternFill("solid", start_color="3A3A1A")
    else:
        cell.fill = PatternFill("solid", start_color="4A1A1A")

    wb.save(str(DATA_FILE))
    print(f"  [XLSX] Logged to {DATA_FILE}")

# ─────────────────────────────────────────────
# SOCIAL MEDIA PUBLISHING
# ─────────────────────────────────────────────

def _post_text(data: dict) -> str:
    """Standard post text used across all platforms."""
    nodes   = data['nodes']
    cycle_s = format_cycle(data.get('cycle_number', 0))
    summary = "  ".join(f"{n['id']}:{n['score']}%" for n in nodes)
    return (
        f"{data['title']}\n"
        f"{cycle_s}\n\n"
        f"COMPOSITE: {data['composite_score']}%\n"
        f"{summary}\n\n"
        f"MASTER: {data['master_node']}\n"
        f"#Interstitial #CycleEngine #GeopoliticalIntelligence"
    )

# ── MASTODON ──────────────────────────────────────────────────────────────────

def publish_to_mastodon(data: dict, img_paths: list, graph_path: Path):
    if not all([MASTODON_ACCESS_TOKEN, MASTODON_INSTANCE]):
        print("  [MASTODON] Credentials not configured. Skipping.")
        return False
    try:
        from mastodon import Mastodon
        m = Mastodon(
            access_token=MASTODON_ACCESS_TOKEN,
            api_base_url=f"https://{MASTODON_INSTANCE}"
        )
        media_ids = []
        for p in [graph_path] + list(img_paths):
            resp = m.media_post(str(p), mime_type="image/jpeg",
                                description=f"Interstitial Cycle Engine node graph for {data['subject']}")
            media_ids.append(resp['id'])
        m.status_post(
            status=_post_text(data)[:500],
            media_ids=media_ids[:4],
            visibility='public'
        )
        print(f"  [MASTODON] Posted: {data['title']}")
        return True
    except Exception as e:
        print(f"  [MASTODON] Error: {e}")
        return False

# ── BLUESKY ───────────────────────────────────────────────────────────────────

def publish_to_bluesky(data: dict, img_paths: list, graph_path: Path):
    if not all([BSKY_HANDLE, BSKY_APP_PASSWORD]):
        print("  [BLUESKY] Credentials not configured. Skipping.")
        return False
    try:
        import requests as req

        # Authenticate
        auth_r = req.post(
            "https://bsky.social/xrpc/com.atproto.server.createSession",
            json={"identifier": BSKY_HANDLE, "password": BSKY_APP_PASSWORD},
            timeout=15
        )
        auth_r.raise_for_status()
        session   = auth_r.json()
        did       = session['did']
        jwt_token = session['accessJwt']
        headers   = {"Authorization": f"Bearer {jwt_token}",
                     "Content-Type": "application/json"}

        # Upload images as blobs (max 4, each max 1MB — resize if needed)
        blob_refs = []
        for p in [graph_path] + list(img_paths):
            img_bytes = open(str(p), 'rb').read()
            # Compress if over 900KB
            if len(img_bytes) > 900_000:
                from PIL import Image as PILImage
                import io
                im = PILImage.open(str(p))
                buf = io.BytesIO()
                im.save(buf, format='JPEG', quality=70)
                img_bytes = buf.getvalue()
            blob_r = req.post(
                "https://bsky.social/xrpc/com.atproto.repo.uploadBlob",
                headers={"Authorization": f"Bearer {jwt_token}",
                         "Content-Type": "image/jpeg"},
                data=img_bytes, timeout=30
            )
            blob_r.raise_for_status()
            blob_refs.append(blob_r.json()['blob'])
            if len(blob_refs) >= 4:
                break

        post_text = _post_text(data)[:300]
        embed = {
            "$type": "app.bsky.embed.images",
            "images": [{"alt": f"ICE node graph {data['subject']}", "image": b}
                       for b in blob_refs]
        }
        record = {
            "$type":   "app.bsky.feed.post",
            "text":    post_text,
            "embed":   embed,
            "createdAt": datetime.datetime.utcnow().strftime('%Y-%m-%dT%H:%M:%SZ'),
            "langs":   ["en"]
        }
        post_r = req.post(
            "https://bsky.social/xrpc/com.atproto.repo.createRecord",
            headers=headers,
            json={"repo": did, "collection": "app.bsky.feed.post", "record": record},
            timeout=30
        )
        post_r.raise_for_status()
        print(f"  [BLUESKY] Posted: {data['title']}")
        return True
    except Exception as e:
        print(f"  [BLUESKY] Error: {e}")
        return False

# ── TELEGRAM ──────────────────────────────────────────────────────────────────

def publish_to_telegram(data: dict, img_paths: list, graph_path: Path):
    if not all([TELEGRAM_BOT_TOKEN, TELEGRAM_CHANNEL_ID]):
        print("  [TELEGRAM] Credentials not configured. Skipping.")
        return False
    try:
        import requests as req

        base = f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}"
        caption = _post_text(data)[:1024]

        # Send as media group (up to 10 images)
        all_imgs = [graph_path] + list(img_paths)
        media = []
        files = {}
        for i, p in enumerate(all_imgs[:4]):
            key = f"photo{i}"
            files[key] = (Path(p).name, open(str(p), 'rb'), 'image/jpeg')
            entry = {"type": "photo", "media": f"attach://{key}"}
            if i == 0:
                entry["caption"]    = caption
                entry["parse_mode"] = "HTML"
            media.append(entry)

        r = req.post(
            f"{base}/sendMediaGroup",
            data={"chat_id": TELEGRAM_CHANNEL_ID, "media": json.dumps(media)},
            files=files,
            timeout=60
        )
        for fobj in files.values():
            fobj[1].close()

        if r.status_code == 200:
            print(f"  [TELEGRAM] Posted: {data['title']}")
            return True
        else:
            print(f"  [TELEGRAM] Error {r.status_code}: {r.text[:200]}")
            return False
    except Exception as e:
        print(f"  [TELEGRAM] Error: {e}")
        return False

# ── LINKEDIN ──────────────────────────────────────────────────────────────────

def publish_to_linkedin(data: dict, img_paths: list, graph_path: Path):
    if not all([LINKEDIN_ACCESS_TOKEN, LINKEDIN_PERSON_URN]):
        print("  [LINKEDIN] Credentials not configured. Skipping.")
        return False
    try:
        import requests as req

        headers = {
            "Authorization": f"Bearer {LINKEDIN_ACCESS_TOKEN}",
            "Content-Type":  "application/json",
            "X-Restli-Protocol-Version": "2.0.0"
        }

        # Register + upload each image
        asset_urns = []
        for p in [graph_path] + list(img_paths):
            # Register upload
            reg_r = req.post(
                "https://api.linkedin.com/v2/assets?action=registerUpload",
                headers=headers,
                json={
                    "registerUploadRequest": {
                        "recipes": ["urn:li:digitalmediaRecipe:feedshare-image"],
                        "owner": LINKEDIN_PERSON_URN,
                        "serviceRelationships": [{
                            "relationshipType": "OWNER",
                            "identifier": "urn:li:userGeneratedContent"
                        }]
                    }
                },
                timeout=20
            )
            reg_r.raise_for_status()
            reg_data   = reg_r.json()
            upload_url = reg_data['value']['uploadMechanism']                         ['com.linkedin.digitalmedia.uploading.MediaUploadHttpRequest']                         ['uploadUrl']
            asset_urn  = reg_data['value']['asset']

            # Upload binary
            with open(str(p), 'rb') as f:
                up_r = req.put(upload_url, data=f,
                               headers={"Authorization": f"Bearer {LINKEDIN_ACCESS_TOKEN}"},
                               timeout=60)
            if up_r.status_code not in (200, 201):
                continue
            asset_urns.append(asset_urn)
            if len(asset_urns) >= 9:
                break

        if not asset_urns:
            print("  [LINKEDIN] No images uploaded successfully.")
            return False

        post_text = _post_text(data)[:3000]
        media_content = [{"status": "READY", "media": u} for u in asset_urns]
        share_content = {
            "shareCommentary": {"text": post_text},
            "shareMediaCategory": "IMAGE",
            "media": media_content
        }
        share_r = req.post(
            "https://api.linkedin.com/v2/ugcPosts",
            headers=headers,
            json={
                "author":          LINKEDIN_PERSON_URN,
                "lifecycleState":  "PUBLISHED",
                "specificContent": {"com.linkedin.ugc.ShareContent": share_content},
                "visibility":      {"com.linkedin.ugc.MemberNetworkVisibility": "PUBLIC"}
            },
            timeout=30
        )
        share_r.raise_for_status()
        print(f"  [LINKEDIN] Posted: {data['title']}")
        return True
    except Exception as e:
        print(f"  [LINKEDIN] Error: {e}")
        return False

# ─────────────────────────────────────────────
# SUBJECT SELECTION — weighted by recency (avoids repetition)
# ─────────────────────────────────────────────

_used_recently = []

def select_subject() -> dict:
    global _used_recently
    pool = [s for s in SUBJECTS if s['name'] not in _used_recently[-30:]]
    if not pool:
        pool = SUBJECTS
    subject = random.choice(pool)
    _used_recently.append(subject['name'])
    if len(_used_recently) > 60:
        _used_recently = _used_recently[-60:]
    return subject

# ─────────────────────────────────────────────
# MAIN RUN CYCLE
# ─────────────────────────────────────────────

def run_cycle():
    now = datetime.datetime.utcnow()
    subject = select_subject()
    cycle_n = increment_cycle()
    cycle_str = format_cycle(cycle_n)
    print(f"\n{'='*60}")
    print(f"[{now.strftime('%Y-%m-%d %H:%M:%S')}] {cycle_str} | SUBJECT: {subject['name']}")
    print(f"{'='*60}")

    # Create run output directory
    safe_name = re.sub(r'[^A-Za-z0-9_]', '_', subject['name'])
    run_dir = OUTPUT_DIR / f"{str(cycle_n).zfill(6)}_{now.strftime('%Y%m%d_%H%M')}_{safe_name}"
    run_dir.mkdir(parents=True, exist_ok=True)

    # Algorithm 1: load refinement context with decay + reset logic
    refinement_ctx = load_refinement_context(subject['name'])

    # Generate report data
    print("  [CLAUDE] Generating cycle report...")
    data = generate_gap_report(subject, refinement_ctx)
    data['cycle_number'] = cycle_n

    # Save raw JSON
    with open(run_dir / "data.json", 'w') as f:
        json.dump(data, f, indent=2)
    print(f"  [JSON] Saved to {run_dir}/data.json")

    # Render page images
    print("  [RENDER] Building page images...")
    img_paths = render_report_images(data, run_dir)

    # Render Markov graph
    print("  [RENDER] Building Markov graph...")
    graph_path = render_markov_graph(data, run_dir)

    # Log to spreadsheet
    print("  [XLSX] Logging run data...")
    log_to_spreadsheet(data, run_dir)

    # Algorithm 2: run diagnostic every 30 cycles
    if cycle_n % 30 == 0:
        print("  [DIAGNOSTIC] Running Algorithm 2 diagnostic...")
        diag = run_diagnostic(cycle_n, data)
        health = diag.get("health", "UNKNOWN")
        print(f"  [DIAGNOSTIC] Health: {health}")
        for w in diag.get("warnings", []):
            print(f"  [DIAGNOSTIC] WARNING: {w[:100]}")
        print(f"  [DIAGNOSTIC] {diag.get('recommendation', '')[:120]}")

    # Publish to social
    print("  [SOCIAL] Publishing...")
    publish_to_mastodon(data, img_paths, graph_path)
    publish_to_bluesky(data, img_paths, graph_path)
    publish_to_telegram(data, img_paths, graph_path)
    publish_to_linkedin(data, img_paths, graph_path)

    print(f"  [DONE] {cycle_str} | {data['title']}")
    print(f"  COMPOSITE: {data['composite_score']}% | MASTER NODE: {data['master_node']}")
    return data

# ─────────────────────────────────────────────
# SCHEDULER
# ─────────────────────────────────────────────

def run_scheduler(interval_minutes: int = 10, run_once: bool = False):
    """Run on schedule. Set run_once=True for single execution."""
    if run_once:
        run_cycle()
        return

    import schedule
    print(f"INTERSTITIAL CYCLE ENGINE ACTIVE // INTERVAL: {interval_minutes} MINUTES")
    print(f"OUTPUT: {OUTPUT_DIR.resolve()}")
    print(f"DATA:   {DATA_FILE.resolve()}")
    print("Press Ctrl+C to stop.\n")

    # Initialise engineering store
    init_engineering_store()
    # Run immediately on start
    run_cycle()

    schedule.every(interval_minutes).minutes.do(run_cycle)

    while True:
        try:
            schedule.run_pending()
            time.sleep(60)
        except KeyboardInterrupt:
            print("\nScheduler stopped.")
            break
        except Exception as e:
            print(f"[ERROR] Cycle failed: {e}")
            traceback.print_exc()
            time.sleep(300)  # Wait 5 min on error before retry

# ─────────────────────────────────────────────
# ENTRY POINT
# ─────────────────────────────────────────────

if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser(description="INTERSTITIAL CYCLE ENGINE")
    parser.add_argument("--once",     action="store_true", help="Run single cycle and exit")
    parser.add_argument("--interval", type=int, default=10, help="Schedule interval in minutes")
    parser.add_argument("--subject",  type=str, default="",  help="Force specific subject name")
    args = parser.parse_args()

    if args.subject:
        match = next((s for s in SUBJECTS if args.subject.lower() in s['name'].lower()), None)
        if match:
            SUBJECTS.insert(0, match)
            _used_recently = [s['name'] for s in SUBJECTS[1:]]
        else:
            print(f"Subject '{args.subject}' not found. Running random selection.")

    run_scheduler(interval_minutes=args.interval, run_once=args.once)
